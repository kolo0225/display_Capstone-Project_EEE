# -*- coding: utf-8 -*-
"""
Created on Sat Apr  3 09:08:33 2021

@author: Kolovos
"""
'''
# --------------------------Assumptions------------------------------
1. GIVES A YEAR FORCAST (main reason missing gas frwd curve in farther future)
 . Also the 'Market values' & 'weather patterns' in M2 will have to be for multiple years 
 
'''

'''
------'Key' for numbering/labeling additional resources and combinations-------

Additional resources
1 = Wind 
2 = Solar
3 = Battery Wind
4 = Battery Solar
5 = Combine Cycle (CC)
6 = Combination Tutbine (CT)
7 = RIC
8 = Aeroderivative (A)
2 conmbinations= double #
3 conmbinations= triple #    (not used in senarios here)
4 conmbinations= quadruple # (not used in senarios here)
'''



# &&&&&&&&&&&&&&&&&&&&& packages &&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&
import numpy as np
import pandas as pd
from pandas import DataFrame
import openpyxl
#from openpyxl.utils import get_column_letter
from openpyxl.styles import Font                       # style page
from openpyxl.styles.borders import Border,Side        # Border
import copy                                            # Wrap, Center..
from openpyxl.chart import BarChart, Reference, Series # excel graph
import matplotlib.pyplot as plt 



############################################################################################################
#                                                                                                          #
# Module 1: Forecast Generation                               updated 3/29/21                              #
#                                                                                                          #
#                                                                                                          #
#     Installed capacity using the year 2021 as based                                                      #
#     of an array of 8760 rows with existing capacity in MWh                                               #
#                                                                                                          #
############################################################################################################

import GUI_values as GUI

#############################################################################################################
# calling values from GUI_values module                                                                     #
# & preparing data                                                                                          #
#                                                                                                           #
#############################################################################################################

location_option = GUI.location                        # value of location to simulate from GUI_values (GUI)

# list of locations available  to simulate
options = ['carousel', 'san_isabel', 'twin_buttes', 'colorado_highlands_wind', 'crossing_trails_wind_farm',
           'kit_carson_wind_farm', 'new_ts_resource']

for i in range(len(options)):    
    if location_option[0] == options[i]:
        loc = options.index(options[i])


utilization = GUI.utilization                           # utilization of transmission line to simulate value from GUI_values (GUI)

print('Utilization from GUI', utilization)
print('location from GUI', location_option)

#################################################################################################################
#                                                                                                               #
#  Bigining of module 1                                                                                         #
#################################################################################################################

# function to convert values from csv file to numbers
def fix_num(num):
    if isinstance(num,str):
        return int(num.replace(',',''))
    else:
        return num


# Reading csv file for LMP8760 & gas8760
market_gas = pd.read_csv('LMP8760_gas8760.csv', usecols = [0,1])
pd.set_option('colheader_justify','center')
#print(market_gas.head(10))

 
# transmission line location to simulate    
if loc == 0:
    

    # reads 8760 csv file
    carousel = pd.read_csv("carousel8760.csv", usecols = [0,1,2,3,4,5,6,7], index_col = ['time'])   # reads a csv file
    pd.set_option('colheader_justify','center')
    #print(carousel.head(10))
    
    carousel.nameplateMWh = carousel.nameplateMWh.apply(fix_num)
    carousel.windMWh = carousel.windMWh.apply(fix_num)
     
    solar = pd.DataFrame(np.array(carousel.solarMWh.apply(fix_num)))
    solar.index = np.arange(1,len(solar)+1)
    solar.columns = ['solar']
          
    lmp = pd.DataFrame(np.array(market_gas.LMP8760.apply(fix_num)))
    lmp.index = np.arange(1,len(lmp)+1)
    lmp.columns = ['LMP']
    
    gas = pd.DataFrame(np.array(market_gas.gas8760.apply(fix_num)))
    gas.index = np.arange(1,len(gas)+1)
    gas.columns = ['gas']
      
    year_base = pd.DataFrame(np.array(carousel.year.apply(fix_num)))
    year_base.index = np.arange(1,len(year_base)+1)
    year_base.columns = ['year']
       
    new_transmission = pd.DataFrame(np.array(carousel.nameplateMWh * utilization),  columns=['new_transmission'])
    new_transmission.index = np.arange(1,len(new_transmission)+1)

    wind = pd.DataFrame(np.array(carousel.windMWh))
    wind.index = np.arange(1,len(wind)+1)
    wind.columns = ['wind']

    forecast = pd.DataFrame(np.array(new_transmission.new_transmission - wind.wind)) 
    forecast.index = np.arange(1,len(forecast)+1)
    forecast.columns = ['forecast']

    full_forecast = pd.DataFrame(np.array(carousel.nameplateMWh - wind.wind))
    full_forecast.index = np.arange(1,len(full_forecast)+1)
    full_forecast.columns = ['full_forecast']
    

    result = pd.concat([year_base,carousel.nameplateMWh,round(full_forecast,2),round(new_transmission,2),round(wind,2),round(forecast,2),round(solar,2),round(lmp,2), round(gas,2)], axis = 1)
    result.columns = ['year','full_transmission','full_forecast','new_transmission','Existing','forecast','Additional','LMP','gas']
    pd.set_option('colheader_justify','center')
    print('')
    print('Carousel forecast')
    print(result.head(10))
        
        
if loc == 1:
    
    
    # reads 8760 csv file
    san_isabel = pd.read_csv("san_isabel8760.csv", usecols = [0,1,2,3,4,5,6,7,8], index_col = ['time'])   # reads a csv file
    pd.set_option('colheader_justify','center')
      

    san_isabel.nameplateMWh = san_isabel.nameplateMWh.apply(fix_num)
    san_isabel.solarMWh = san_isabel.solarMWh.apply(fix_num)
   
    
    solar = pd.DataFrame(np.array(san_isabel.solarMWh.apply(fix_num)))
    solar.index = np.arange(1,len(solar)+1)
    solar.columns = ['solar']
    
        
    lmp = pd.DataFrame(np.array(market_gas.LMP8760.apply(fix_num)))
    lmp.index = np.arange(1,len(lmp)+1)
    lmp.columns = ['LMP']
    
    gas = pd.DataFrame(np.array(market_gas.gas8760.apply(fix_num)))
    gas.index = np.arange(1,len(gas)+1)
    gas.columns = ['gas']
    
    
    year_base = pd.DataFrame(np.array(san_isabel.year.apply(fix_num)))
    year_base.index = np.arange(1,len(year_base)+1)
    year_base.columns = ['year']
       
    new_transmission = pd.DataFrame(np.array(san_isabel.nameplateMWh * utilization),  columns=['new_transmission'])
    new_transmission.index = np.arange(1,len(new_transmission)+1)

    wind = pd.DataFrame(np.array(san_isabel.windMWh.apply(fix_num)))
    wind.index = np.arange(1,len(wind)+1)
    wind.columns = ['wind']

    forecast = pd.DataFrame(np.array(new_transmission.new_transmission - solar.solar))
    forecast.index = np.arange(1,len(forecast)+1)
    forecast.columns = ['forecast']

    full_forecast = pd.DataFrame(np.array(san_isabel.nameplateMWh - solar.solar))
    full_forecast.index = np.arange(1,len(full_forecast)+1)
    full_forecast.columns = ['full_forecast']
    

    result = pd.concat([year_base,san_isabel.nameplateMWh,round(full_forecast,2),round(new_transmission,2),round(solar,2),round(forecast,2),round(wind,3),round(lmp,2), round(gas,2)], axis = 1)
    result.columns = ['year','full_transmission','full_forecast','new_transmission','Existing','forecast','Additional','LMP','gas']
    pd.set_option('colheader_justify','center')
    print('')
    print('San Isabel forecast')
    print(result.head(10))
    

if loc == 2:
    
    
    # reads 8760 csv file
    twin_buttes = pd.read_csv("twin_buttes8760.csv", usecols = [0,1,2,3,4,5,6,7,8], index_col = ['time'])   # reads a csv file
    pd.set_option('colheader_justify','center')
    

    twin_buttes.nameplateMWh = twin_buttes.nameplateMWh.apply(fix_num)
    twin_buttes.windMWh = twin_buttes.windMWh.apply(fix_num)
     
    solar = pd.DataFrame(np.array(twin_buttes.solarMWh.apply(fix_num)))
    solar.index = np.arange(1,len(solar)+1)
    solar.columns = ['solar']
          
    lmp = pd.DataFrame(np.array(market_gas.LMP8760.apply(fix_num)))
    lmp.index = np.arange(1,len(lmp)+1)
    lmp.columns = ['LMP']
    
    gas = pd.DataFrame(np.array(market_gas.gas8760.apply(fix_num)))
    gas.index = np.arange(1,len(gas)+1)
    gas.columns = ['gas']
      
    year_base = pd.DataFrame(np.array(twin_buttes.year.apply(fix_num)))
    year_base.index = np.arange(1,len(year_base)+1)
    year_base.columns = ['year']
       
    new_transmission = pd.DataFrame(np.array(twin_buttes.nameplateMWh * utilization),  columns=['new_transmission'])
    new_transmission.index = np.arange(1,len(new_transmission)+1)

    wind = pd.DataFrame(np.array(twin_buttes.windMWh))
    wind.index = np.arange(1,len(wind)+1)
    wind.columns = ['wind']

    forecast = pd.DataFrame(np.array(new_transmission.new_transmission - wind.wind))
    forecast.index = np.arange(1,len(forecast)+1)
    forecast.columns = ['forecast']

    full_forecast = pd.DataFrame(np.array(twin_buttes.nameplateMWh - wind.wind))
    full_forecast.index = np.arange(1,len(full_forecast)+1)
    full_forecast.columns = ['full_forecast']
    

    result = pd.concat([year_base,twin_buttes.nameplateMWh,round(full_forecast,2),round(new_transmission,2),round(wind,2),round(forecast,2),round(solar,2),round(lmp,2), round(gas,2)], axis = 1)
    result.columns = ['year','full_transmission','full_forecast','new_transmission','Existing','forecast','Additional','LMP','gas']
    pd.set_option('colheader_justify','center')
    print('')
    print('Twin Buttes forecast')
    print(result.head(10))

    


if loc == 3:
    
    
    # reads 8760 csv file
    colorado_highlands_wind = pd.read_csv("colorado_highlands8760.csv", usecols = [0,1,2,3,4,5,6,7,8], index_col = ['time'])   # reads a csv file
    pd.set_option('colheader_justify','center')
    
    
    colorado_highlands_wind.nameplateMWh = colorado_highlands_wind.nameplateMWh.apply(fix_num)
    colorado_highlands_wind.windMWh = colorado_highlands_wind.windMWh.apply(fix_num)
     
    solar = pd.DataFrame(np.array(colorado_highlands_wind.solarMWh.apply(fix_num)))
    solar.index = np.arange(1,len(solar)+1)
    solar.columns = ['solar']
          
    lmp = pd.DataFrame(np.array(market_gas.LMP8760.apply(fix_num)))
    lmp.index = np.arange(1,len(lmp)+1)
    lmp.columns = ['LMP']
    
    gas = pd.DataFrame(np.array(market_gas.gas8760.apply(fix_num)))
    gas.index = np.arange(1,len(gas)+1)
    gas.columns = ['gas']
      
    year_base = pd.DataFrame(np.array(colorado_highlands_wind.year.apply(fix_num)))
    year_base.index = np.arange(1,len(year_base)+1)
    year_base.columns = ['year']
       
    new_transmission = pd.DataFrame(np.array(colorado_highlands_wind.nameplateMWh * utilization),  columns=['new_transmission'])
    new_transmission.index = np.arange(1,len(new_transmission)+1)

    wind = pd.DataFrame(np.array(colorado_highlands_wind.windMWh))
    wind.index = np.arange(1,len(wind)+1)
    wind.columns = ['wind']

    forecast = pd.DataFrame(np.array(new_transmission.new_transmission - wind.wind))
    forecast.index = np.arange(1,len(forecast)+1)
    forecast.columns = ['forecast']

    full_forecast = pd.DataFrame(np.array(colorado_highlands_wind.nameplateMWh - wind.wind))
    full_forecast.index = np.arange(1,len(full_forecast)+1)
    full_forecast.columns = ['full_forecast']
    

    result = pd.concat([year_base,colorado_highlands_wind.nameplateMWh,round(full_forecast,2),round(new_transmission,2),round(wind,2),round(forecast,2),round(solar,2),round(lmp,2), round(gas,2)], axis = 1)
    result.columns = ['year','full_transmission','full_forecast','new_transmission','Existing','forecast','Additional','LMP','gas']
    pd.set_option('colheader_justify','center')
    print('')
    print('Colorado Highlands Wind forecast')
    print(result.head(10))

    
    
if loc == 4:
    
    
    # reads 8760 csv file
    crossing_trails_wind_farm = pd.read_csv("crossing_trails_wind_farm8760.csv", usecols = [0,1,2,3,4,5,6,7,8], index_col = ['time'])   # reads a csv file
    pd.set_option('colheader_justify','center')
    
    
    crossing_trails_wind_farm.nameplateMWh = crossing_trails_wind_farm.nameplateMWh.apply(fix_num)
    crossing_trails_wind_farm.windMWh = crossing_trails_wind_farm.windMWh.apply(fix_num)
     
    solar = pd.DataFrame(np.array(crossing_trails_wind_farm.solarMWh.apply(fix_num)))
    solar.index = np.arange(1,len(solar)+1)
    solar.columns = ['solar']
          
    lmp = pd.DataFrame(np.array(market_gas.LMP8760.apply(fix_num)))
    lmp.index = np.arange(1,len(lmp)+1)
    lmp.columns = ['LMP']
    
    gas = pd.DataFrame(np.array(market_gas.gas8760.apply(fix_num)))
    gas.index = np.arange(1,len(gas)+1)
    gas.columns = ['gas']
      
    year_base = pd.DataFrame(np.array(crossing_trails_wind_farm.year.apply(fix_num)))
    year_base.index = np.arange(1,len(year_base)+1)
    year_base.columns = ['year']
       
    new_transmission = pd.DataFrame(np.array(crossing_trails_wind_farm.nameplateMWh * utilization),  columns=['new_transmission'])
    new_transmission.index = np.arange(1,len(new_transmission)+1)

    wind = pd.DataFrame(np.array(crossing_trails_wind_farm.windMWh))
    wind.index = np.arange(1,len(wind)+1)
    wind.columns = ['wind']

    forecast = pd.DataFrame(np.array(new_transmission.new_transmission - wind.wind))
    forecast.index = np.arange(1,len(forecast)+1)
    forecast.columns = ['forecast']

    full_forecast = pd.DataFrame(np.array(crossing_trails_wind_farm.nameplateMWh - wind.wind))
    full_forecast.index = np.arange(1,len(full_forecast)+1)
    full_forecast.columns = ['full_forecast']
    

    result = pd.concat([year_base,crossing_trails_wind_farm.nameplateMWh,round(full_forecast,2),round(new_transmission,2),round(wind,2),round(forecast,2),round(solar,2),round(lmp,2), round(gas,2)], axis = 1)
    result.columns = ['year','full_transmission','full_forecast','new_transmission','Existing','forecast','Additional','LMP','gas']
    pd.set_option('colheader_justify','center')
    print('')
    print('Crossing Trails Wind Farm forecast')
    print(result.head(10))

    

if loc == 5:
    
    
    # reads 8760 csv file
    kit_carson_wind_farm = pd.read_csv("kit_carson_wind_farm8760.csv", usecols = [0,1,2,3,4,5,6,7,8], index_col = ['time'])   # reads a csv file
    pd.set_option('colheader_justify','center')
    

    kit_carson_wind_farm.nameplateMWh = kit_carson_wind_farm.nameplateMWh.apply(fix_num)
    kit_carson_wind_farm.windMWh = kit_carson_wind_farm.windMWh.apply(fix_num)
     
    solar = pd.DataFrame(np.array(kit_carson_wind_farm.solarMWh.apply(fix_num)))
    solar.index = np.arange(1,len(solar)+1)
    solar.columns = ['solar']
          
    lmp = pd.DataFrame(np.array(market_gas.LMP8760.apply(fix_num)))
    lmp.index = np.arange(1,len(lmp)+1)
    lmp.columns = ['LMP']
    
    gas = pd.DataFrame(np.array(market_gas.gas8760.apply(fix_num)))
    gas.index = np.arange(1,len(gas)+1)
    gas.columns = ['gas']
      
    year_base = pd.DataFrame(np.array(kit_carson_wind_farm.year.apply(fix_num)))
    year_base.index = np.arange(1,len(year_base)+1)
    year_base.columns = ['year']
       
    new_transmission = pd.DataFrame(np.array(kit_carson_wind_farm.nameplateMWh * utilization),  columns=['new_transmission'])
    new_transmission.index = np.arange(1,len(new_transmission)+1)

    wind = pd.DataFrame(np.array(kit_carson_wind_farm.windMWh))
    wind.index = np.arange(1,len(wind)+1)
    wind.columns = ['wind']

    forecast = pd.DataFrame(np.array(new_transmission.new_transmission - wind.wind))
    forecast.index = np.arange(1,len(forecast)+1)
    forecast.columns = ['forecast']

    full_forecast = pd.DataFrame(np.array(kit_carson_wind_farm.nameplateMWh - wind.wind))
    full_forecast.index = np.arange(1,len(full_forecast)+1)
    full_forecast.columns = ['full_forecast']
    

    result = pd.concat([year_base,kit_carson_wind_farm.nameplateMWh,round(full_forecast,2),round(new_transmission,2),round(wind,2),round(forecast,2),round(solar,2),round(lmp,2), round(gas,2)], axis = 1)
    result.columns = ['year','full_transmission','full_forecast','new_transmission','Existing','forecast','Additional','LMP','gas']
    pd.set_option('colheader_justify','center')
    print('')
    print('Kit Carson Wind Farm forecast')
    print(result.head(10))

    

if loc == 6:
    
    
    # reads 8760 csv file
    new_ts_resource = pd.read_csv("new_ts_resource8760.csv", usecols = [0,1,2,3,4,5,6,7,8], index_col = ['time'])   # reads a csv file
    pd.set_option('colheader_justify','center')
    

    new_ts_resource.nameplateMWh = new_ts_resource.nameplateMWh.apply(fix_num)
    new_ts_resource.windMWh = new_ts_resource.windMWh.apply(fix_num)
     
    solar = pd.DataFrame(np.array(new_ts_resource.solarMWh.apply(fix_num)))
    solar.index = np.arange(1,len(solar)+1)
    solar.columns = ['solar']
          
    lmp = pd.DataFrame(np.array(market_gas.LMP8760.apply(fix_num)))
    lmp.index = np.arange(1,len(lmp)+1)
    lmp.columns = ['LMP']
    
    gas = pd.DataFrame(np.array(market_gas.gas8760.apply(fix_num)))
    gas.index = np.arange(1,len(gas)+1)
    gas.columns = ['gas']
      
    year_base = pd.DataFrame(np.array(new_ts_resource.year.apply(fix_num)))
    year_base.index = np.arange(1,len(year_base)+1)
    year_base.columns = ['year']
       
    new_transmission = pd.DataFrame(np.array(new_ts_resource.nameplateMWh * utilization),  columns=['new_transmission'])
    new_transmission.index = np.arange(1,len(new_transmission)+1)

    wind = pd.DataFrame(np.array(new_ts_resource.windMWh))
    wind.index = np.arange(1,len(wind)+1)
    wind.columns = ['wind']

    forecast = pd.DataFrame(np.array(new_transmission.new_transmission - wind.wind))
    forecast.index = np.arange(1,len(forecast)+1)
    forecast.columns = ['forecast']

    full_forecast = pd.DataFrame(np.array(new_ts_resource.nameplateMWh - wind.wind))
    full_forecast.index = np.arange(1,len(full_forecast)+1)
    full_forecast.columns = ['full_forecast']
    

    result = pd.concat([year_base,new_ts_resource.nameplateMWh,round(full_forecast,2),round(new_transmission,2),round(wind,2),round(forecast,2),round(solar,2),round(lmp,2), round(gas,2)], axis = 1)
    result.columns = ['year','full_transmission','full_forecast','new_transmission','Existing','forecast','Additional','LMP','gas']
    pd.set_option('colheader_justify','center')
    print('')
    print('New TS Resource forecast')
    print(result.head(10))

    
# ########################## End Module1  ###################################

# @@@@@@@@@@@@@@@@@@@@@@@ MODULE 2 @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
# @@@@@@@@@@@@@@@@@@@@@ data retrieves @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
# ================ data_retrieve for Module 1==================================
# Picks values from results Module1, creates an: 
#                        1. list 8760, 
#                        2. sums the list
#                        3. picks up 1st value of list (used for constants)
#   input:
#       result : from the result of Module1 it picks values
#   output:
#       list_8760
#       Sum       : sums the list
#       constant  : picks up 1st value of list (used for constant)   
# ['year','full_transmission','full_forecast','new_transmission',
# 'Existing','forecast','Additional','LMP','gas']
def data_retrieve (result): 
    list_8760 =[ ]
    Sum = 0 
    constant = 0             
    for i in result:
        list_8760.append(i)
        Sum += i
        constant =  list_8760[0]
    return list_8760, Sum, constant

nameplate8760, Sum_nameplate, nameplate = data_retrieve (result['full_transmission'])
new_trans8760, Sum_new_trans, new_trans = data_retrieve (result['new_transmission'])
Existing8760, Sum_Existing, ignore0 = data_retrieve (result['Existing'])
Additional8760, ignore1, ignore2 = data_retrieve (result['Additional'])
full_forecast8760, Sum_full_forecast, ignore3 = data_retrieve (result['full_forecast'])
forecast8760, Sum_forecast, ignore4 = data_retrieve (result['forecast'])

gas8760, ignore5, ignore6 = data_retrieve (result['gas'])
LMP8760, ignore7, ignore8  = data_retrieve (result['LMP'])


#------------------- Data retrieved using pandas ------------------
# Reading data from 'EconomicAnalysisM3.xlsx' in 'project program' folder
# pandas can read formulas directly from excel sheet

sheet = pd.read_excel('EconomicAnalysisM3.xlsx') 

# ------------- picking up values from xlsx using pandas ----------------
# Setting the constants from xlsx
# input:
#    xl1: row
#    xl2: column
# output
#   xlsx : value form xlsx cell
def pickFROMxlsx (xl1, xl2):
    xlsx = sheet.iloc[xl1, xl2]   # [BTU/kWh] 
    return xlsx

# CC, F9, F7 of 'EconomicAnalysisM3.xlsx'
CC_HR = pickFROMxlsx (7, 5)
CC_VOM = pickFROMxlsx (5, 5)
# CT, G9, G7 of 'EconomicAnalysisM3.xlsx'
CT_HR = pickFROMxlsx (7, 6)
CT_VOM = pickFROMxlsx (5, 6)
# RIC, H9, H7 of 'EconomicAnalysisM3.xlsx'
RIC_HR = pickFROMxlsx (7, 7)
RIC_VOM = pickFROMxlsx (5, 7)
# ARO, I9, I7 of 'EconomicAnalysisM3.xlsx' 
ARO_HR = pickFROMxlsx (7, 8)
ARO_VOM = pickFROMxlsx (5, 8)


#print("ARO_HR: ",ARO_HR, "\n","ARO_VOM: ", ARO_VOM )
# @@@@@@@@@@@@@@@@@@@@@ end of data retrives @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

# @@@@@@@@@@@@@@@@@@@@ Control panel M2 @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

#============= Control buttons constants and functions ====================
# used to determine sizes (ex: 0,5,10,15,20............)
# End:
#   carousel, (loc=0) = 300
#   carousel, (loc=1) = 600
START = 0                 # size of smallest farm tested, 0 is recommended
END = 2 * nameplate       # size of largest farm tested, [MWh] // 2*nameplate
N = 60                    # total number of steps

# --------------- STEP size ------------------------------------------------ 
          
STEP = int((END-START)/N)  # Size of farms increases by [MWh]
print ("STEP: ", STEP) 

##################################################################################################      
# Gui valuse for module 2:                                                                           #
##################################################################################################
#                      ideal Resources picked by user to test
# Renewable resources to simulate           
aro = GUI.aro_value      
batteries = GUI.batteries_value   
CC = GUI.CC_value 
CT = GUI.CT_value   
RIC = GUI.RIC_value   
solar = GUI.solar_value   
wind = GUI.wind_value


selection = ['No', 'Yes'] 

# -------------- No = 0, Yes =1 ---------------------------------------------
#   input:
#       element: 0 or 1 position in the list ( needed for size array later)
#       resource: wind, solar........ 
#       selection: list['str1', 'str2' for asssign position int to list] 
#   output:
#       choise : 0 if No, 1 if yes (because 'Yes' is in position 1 of the list)
def str_in_list_to_int (element, resource, selection):
    for i in range(len(selection)):  
        if resource[element] == selection[i]:
            choise = i
    return choise

# values of 0 or 1 are assigned to ideal resources (to be tested or not)
a_ = str_in_list_to_int (0, wind, selection)
b_ = str_in_list_to_int (0, solar, selection) 
c_ = str_in_list_to_int (0, CC, selection) 
d_ = str_in_list_to_int (0, CT, selection)
e_ = str_in_list_to_int (0, RIC, selection)
f_ = str_in_list_to_int (0, aro, selection)    
g_ = str_in_list_to_int (0, batteries, selection) 

# ------------ Resources 'Ideal values'  user wants to evaluate --------------
# if var chosen then =1 and displaied
wind_Id = a_       # @ M3 Energy & curt through if ...then,  
solar_Id = b_   # keep always 1 regarless GUI, to prevent complications
# choice takes place @ 'best_choice', effects 'monthEnergyGas' fn (single & comb.)
CC_Id = c_
CT_Id = d_
RIC_Id = e_
ARO_Id = f_
# choice takes place @ 'size_resource' then @ "energyAdditionalPlusBatteries"
Bat1_Id =Bat2_Id =Bat3_Id = g_  # if batery is chosen (in any) all dispayed

# ------------ Resources 'USER CHOICE values'  user wants to evaluate ---------
# Farm size to simulate  from GUI  
size1 = GUI.size1_value
size2 = GUI.size2_value        
size3 = GUI.size3_value

# list of possible answer given at GUI
selection1_user = ['Choose a size option #1','Additional resource 1', 'CC1', 'CT1', 
                  'RIC1', 'ARO1', 'Battery 1']
selection2_user = ['Choose a size option #2','Additional resource 2', 'CC2', 'CT2', 
                  'RIC2', 'ARO2', 'Battery 2']
selection3_user = ['Choose a size option #3','Battery 3']

# picks user choices to test, as a element number in selection_user[]
# ex user_choice[0,0,1,0,0....] hot code
user_choice1 = str_in_list_to_int (0, size1, selection1_user)
user_choice2 = str_in_list_to_int (0, size2, selection2_user)
user_choice3 = str_in_list_to_int (0, size3, selection3_user)

# -------------- No = 0, Yes =1 ---------------------------------------------
#   input:
#       user_decision: the reource that the user wants to test (as int 0-6)
#       selections: list['str1', 'str2' to assign length of for loop
#   output:
#      variable : variable [0,0,1,0..] hot code, if 1 user choice resource used
def int_assigned_to_var (user_decision, selections):
    variable =[]
    for i in range(len(selections)):  
        
        if user_decision == i:
           x = 1
        else:
            x = 0
        variable.append(x)
    return variable 

# user choice determines the variables that will be tested
variable1 = int_assigned_to_var (user_choice1, selection1_user)
variable2 = int_assigned_to_var (user_choice2, selection2_user)
variable3 = int_assigned_to_var (user_choice3, selection3_user)


# if var chosen then =1 and displaied if not 0 and not dispayed
# @ 'size_resource' @ array_position @ 'Sum_Add_added[nAdd1,2]' fn 
# chs_add1 apply to (single & comb.) so also 'best_choice',chs_add2 only single
chs_add1 = variable1[1]  #  Additional resource chosen size 1 
chs_add2 = variable2[1]  #  Additional resource chosen size 2

# @ 'size_resource' @ array_position @ 'monthEnergyGas' (single & comb.)fn 
chs_CC1 = variable1[2]  # gas resource chosen size 1 
chs_CT1 = variable1[3]  # gas resource chosen size 1 
chs_RIC1 = variable1[4] # gas resource chosen size 1 
chs_ARO1 = variable1[5] # gas resource chosen size 1 

chs_CC2 = variable2[2]  # gas resource chosen size 2
chs_CT2 = variable2[3]  # gas resource chosen size 2
chs_RIC2 = variable2[4] # gas resource chosen size 2
chs_ARO2 = variable2[5] # gas resource chosen size 2

# battery resources chosen size1, 2 & 3
chs_Bat1 = variable1[6] # Bat1_Id & chs_Bat1 have to be chosen to display bat.
chs_Bat2 = variable2[6] # Bat1_Id & chs_Bat1 have to be chosen to display bat.
chs_Bat3 = variable3[1] # Bat1_Id & chs_Bat1 have to be chosen to display bat.

# # turns str 'value' in to a float for determining the size of the teted resource
#   input:
#       size_chosen: the size chosen by the user (fraction of nameplate)
#   output:
#     size_ : it is an float
def str_to_int (size_chosen):
    num = size_chosen[1]
    size_= float(num)
    return size_

# -- user choises sizes as fraction of nameplate for 0 to 1 ------------------
size1_ = str_to_int (size1)
size2_ = str_to_int (size2) 
size3_ = str_to_int (size3)

# -- x = 1s2, y =2nd, z = 3rd ------------------------------------------------
x = size1_
y = size2_
z = size3_

# ------------------ farm size function ---------------------------------
# chooses the size of the farm based on var (x,y,z) & namplate 
#   input:
#       var (x,y,z)
#       napleate [MWh]
#   output:
#       size : size of the farm

def size_resource (var, nameplate, chosen):
    size = int(var*nameplate*chosen)
    return size

SZadd1 = size_resource (x, nameplate,chs_add1 ) # choose size of the farm to test [MWh]
SZadd2 = size_resource (y, nameplate,chs_add2 ) # choose size of the farm to test [MWh]

SZCC1 = size_resource (x, nameplate,chs_CC1)  # choose size of the farm to test [MWh]
SZCC2 = size_resource (y, nameplate,chs_CC2)  # choose size of the farm to test [MWh]
SZCT1 = size_resource (x, nameplate,chs_CT1)  # choose size of the farm to test [MWh]
SZCT2 = size_resource (y, nameplate,chs_CT2)  # choose size of the farm to test [MWh]
SZRIC1 = size_resource (x, nameplate,chs_RIC1) # choose size of the farm to test [MWh]
SZRIC2 = size_resource (y, nameplate,chs_RIC2) # choose size of the farm to test [MWh]
SZARO1 = size_resource (x, nameplate,chs_ARO1) # choose size of the farm to test [MWh]
SZARO2 = size_resource (y, nameplate,chs_ARO2) # choose size of the farm to test [MWh]
# battery user can test up to 3 sizes @ the time
SZbat1 = size_resource (x, nameplate,Bat1_Id and chs_Bat1) # size of the farm to test [MWh]
SZbat2 = size_resource (y, nameplate,Bat2_Id and chs_Bat2) # size of the farm to test [MWh]
SZbat3 = size_resource (z, nameplate,Bat3_Id and chs_Bat3) # size of the farm to test [MWh]

# choose size of the farm want to test [MWh] - combined Additional + solar
#SZaddCC1 = int(.2*nameplate)   # user chooses 'SZadd1'  code chooses SZaddCC1
SZaddCC2 = size_resource (x, nameplate,chs_CC1) # Additional = ideal, gas = choice
SZaddCC3 = size_resource (y, nameplate,chs_CC2) # user chooses 'SZadd1' & SZaddCC3

SZaddCT2 = size_resource (x, nameplate,chs_CT1) # Additional = ideal, gas = choice
SZaddCT3 = size_resource (y, nameplate,chs_CT2) # user chooses 'SZadd1' & SZaddCC3

SZaddRIC2 = size_resource (x, nameplate,chs_RIC1) # Additional = ideal, gas = choice
SZaddRIC3 = size_resource (y, nameplate,chs_RIC2) # user chooses 'SZadd1' & SZaddCC3

SZaddARO2 = size_resource (x, nameplate,chs_ARO1) # Additional = ideal, gas = choice
SZaddARO3 = size_resource (y, nameplate,chs_ARO2) # user chooses 'SZadd1' & SZaddCC3

# ------------------ Array Position function ---------------------------------
# chooses the position in list based on size (SZ) chosen 
#   input:
#       size
#   output:
#       position : position in the array to get desirable SZ of farm
def array_position (size):
    position= int((size-START)/STEP)  # since int() in will prevent it from stop working
    return position
nAdd1 = array_position (SZadd1)         # positions of array based on SZ chosen
nAdd2 = array_position (SZadd2)         # positions of array based on SZ chosen
nCC1 = array_position (SZCC1)           # positions of array based on SZ chosen
nCC2 = array_position (SZCC2)           # positions of array based on SZ 
nCT1 = array_position (SZCT1)           # positions of array based on SZ chosen           
nCT2 = array_position (SZCT2)           # positions of array based on SZ chosen
nRIC1 = array_position (SZRIC1)         # positions of array based on SZ chosen
nRIC2 = array_position (SZRIC2)         # positions of array based on SZ chosen
nARO1 = array_position (SZARO1)         # positions of array based on SZ chosen
nARO2 = array_position (SZARO2)         # positions of array based on SZ chosen

# position : position in the array to get desirable SZ battery
nBat1 = array_position (SZbat1)         # positions of array based on SZ chosen
nBat2 = array_position (SZbat2)         # positions of array based on SZ chosen
nBat3 = array_position (SZbat3)         # positions of array based on SZ chosen

# positions of array based on SZ chosen- combined Additional + solar
#nAddCC1 = array_position (SZaddCC1)     # user- 'SZadd1',  code- SZaddCC1
nAddCC2 = array_position (SZaddCC2)     # Additional = ideal, gas = choice
nAddCC3 = array_position (SZaddCC3)     # user- 'SZadd1',  user- SZaddCC3

nAddCT2 = array_position (SZaddCT2)     # Additional = ideal, gas = choice
nAddCT3 = array_position (SZaddCT3)     # user- 'SZadd1',  user- SZaddCC3

nAddRIC2 = array_position (SZaddRIC2)     # Additional = ideal, gas = choice
nAddRIC3 = array_position (SZaddRIC3)     # user- 'SZadd1',  user- SZaddCC3

nAddARO2 = array_position (SZaddARO2)     # Additional = ideal, gas = choice
nAddARO3 = array_position (SZaddARO3)     # user- 'SZadd1',  user- SZaddCC3

# --------- elements in the list to be tested start in 8760--------------
# used in a 8760 array // | Used for 'print' only |
# leters from a to f can be reserve for positions in array & matrix
a =7980     # used for column position (8760)                 
b= 8000     # used for column position (8760) 
c = 0       # used for row - 'size's position (N) 
d = 10      # used for row - 'size's position (N)                 

# ###########################################################################
# ========================== Control plots buttons in M2 =====================
# use this variables to determine the plot [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
resources = ['Added','CC','CT','RIC','ARO','Add_battery', \
            'Add_CC','Add_CT','Add_RIC','Add_ARO']
    
# ============== choose the graph to plot ===============================
# Plot for farm size to display  
plot = GUI.plot_value

# list with all possible plot to be displayed as determined by q
selection_plot = ['Size vs Size','Energy vs Size','Curtail vs Size','Utilization vs Size',
                  '1st Derivative vs Size','2nd Derivative vs Size'] 

q = str_in_list_to_int (0, plot, selection_plot)
     
# NAMES = [0. sizes, 1. Sum_added, 2. Sum_curtail, 3. utilization_act,\
#          4. FstDeriv, 5. SecDeriv  ]  
# determines the y vs x 
p = 0               #  value is always 0

#  utilize q to choose a plots 
Additional_var_x =p
Additional_var_y=q

CC_var_x =p
CC_var_y=q

CT_var_x =p
CT_var_y=q

RIC_var_x =p
RIC_var_y=q

ARO_var_x =p
ARO_var_y=q

AddBat_var_x =p
AddBat_var_y=q

Add_CC_var_x =p
Add_CC_var_y=q

Add_CT_var_x =p
Add_CT_var_y=q

Add_ARO_var_x =p
Add_ARO_var_y=q

Add_RIC_var_x =p
Add_RIC_var_y=q


##############################################################################

# @@@@@@@@@@@@@@@@@@@@@@ single gas @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
# --------- Converting Gas prices into  [$/MWh] --------------------------- 
# input:
#    gas8760: list 8760 of gas prices in [$/mmBTU]
#    HR: heat rate of gas in [BTU/kWh], 1 var not list
#    VOM:[$/MWh], 1 var not list
# output
#   gasPrice8760 :list 8760 of gas prices in [$/MWh]

def Dol_mmBTUtoDol_MWh (gas8760, HR, VOM):
    # create a empty list for CC, CT, RIC, ARO
    gasPrice8760 = []      # list 8760
    
    # * HR [BTU/MWh] + VOM [$/MWh]
    for i in range(len(gas8760)):
        gasPrice8760.append(((gas8760[i] *HR)/1000) + VOM)      
    return gasPrice8760
CC_gasPrice8760 = Dol_mmBTUtoDol_MWh (gas8760, CC_HR, CC_VOM)
CT_gasPrice8760 = Dol_mmBTUtoDol_MWh (gas8760, CT_HR, CT_VOM)
RIC_gasPrice8760 = Dol_mmBTUtoDol_MWh (gas8760, RIC_HR, RIC_VOM)
ARO_gasPrice8760 = Dol_mmBTUtoDol_MWh (gas8760, ARO_HR, ARO_VOM)

#print("\nCT_gasPrice8760[0:10] : \n", CT_gasPrice8760[c:d])
#print("CT_gasPrice8760 len: ",len(CT_gasPrice8760),"\n")

#print("ARO_HR: ",ARO_HR, "\n","ARO_VOM: ", ARO_VOM )
#-----------------------------------------------------------------------

# -------- comparing gas to market, places a 0 when gasPrice > Market)--------
# ---------Compares listA with listB, places 0 when listA > listB---------
# formula =if gas <= market use gas
# gasPrice = list of gas price (when gasPrice<=Market) & 0 otherwise)
# input:
#    gasPrice8760 :list 8760 of gas prices in [$/MWh]
#    LMP8760 :list 8760 of gas prices in [$/MWh]
# output
#   gasPrice8760 :list 8760 of gas prices in [$/MWh]

def gasVSmarket (gasPrice8760, LMP8760):
    
    for i in range (len(gasPrice8760)): 
        if gasPrice8760[i] <= LMP8760[i]:
            gasPrice8760[i] = gasPrice8760[i]
        else:
           gasPrice8760 [i]= 0
    return gasPrice8760

CC_gasPrice8760 = gasVSmarket (CC_gasPrice8760, LMP8760)
CT_gasPrice8760 = gasVSmarket (CT_gasPrice8760, LMP8760)
RIC_gasPrice8760 = gasVSmarket (RIC_gasPrice8760, LMP8760)
ARO_gasPrice8760 = gasVSmarket (ARO_gasPrice8760, LMP8760)

#print("CT_gasPrice8760[0:10]: \n",CT_gasPrice8760[c:d])
#print("CT_gasPrice8760 len: ",len(CT_gasPrice8760),"\n")

# --------------------------------------------------------------------
# Compares listA with ListB places0 when listB = 0, conAddtant otherwise
# Using gasPrice8760 (now with 0 where > market)
# Creates energyGas8760 based on MHh =100 if gasPrice8760 != 0
# lists that hold all 0's for CC, CT, RIC, ARO, (gasPrice > marketPrice)
# input:
#    MWh : 100 MWh when gasPrice8760 != 0, 0 else
#    gasPrice8760 : compares with energyGas8760  to determine if 0 or 100
# output
#   energyGas8760 :list 8760 of gas energy [MWh]
def gasENERGYprofileLIST (MWh, gasPrice8760):
    energyGas8760 = []
    for i in range(len(gasPrice8760)):
        if gasPrice8760[i] != 0 :
            energyGas8760.append(MWh)  # 100 MWh list 8760
        else :
            energyGas8760.append(0)  # 100 MWh list 8760
    return energyGas8760
CC_energyGas8760 = gasENERGYprofileLIST (100, CC_gasPrice8760 )
CT_energyGas8760 = gasENERGYprofileLIST (100, CT_gasPrice8760 )
RIC_energyGas8760 = gasENERGYprofileLIST (100, RIC_gasPrice8760 )
ARO_energyGas8760 = gasENERGYprofileLIST (100, ARO_gasPrice8760 )
#print("CC_energyGas8760[c:d]: \n",CC_energyGas8760[c:d])
#print("CC_energyGas8760 len: ",len(CC_energyGas8760),"\n")
# --------------------------------------------------------------------
# ---------------------------------------------------------------------------
# --------------- size of energyGas8760 > 0, || not used in program ||---------
# Creates listA from ListB only when listB != 0
# Used to count the indexes that energyGas8760 != 0
# input:
#    energyGas8760 :list 8760 of gas energy [MWh]
# output
#   UseGASn : list with hrs that energyGas8760 will be utilized

def canUSEgas (energyGas8760):

    UseGASn=[]
    for i in range (len(energyGas8760)): 
        if energyGas8760[i] != 0:
            UseGASn.append(i)
    return UseGASn
      
# print lenght of list gasPrice == 0 (gasPrice > marketPrice)
CCn = canUSEgas (CC_gasPrice8760)
CTn = canUSEgas (CT_gasPrice8760)
RICn = canUSEgas (RIC_gasPrice8760)
AROn = canUSEgas (ARO_gasPrice8760)

# @@@@@@@@@@@@@@@@@@@@@@@@@@@@ end gas @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
# ------------------------------------------------------------------------

# ^^^^^^^^^ array of sizes (sizes) of farms [MWh] ^^^^^^^^^^^^^^^^^^^^^^^^ 
# ******** for gas size = energy output [MWh]
# Creates an array of sizes (sizes) of farms function, based on (START,END,STEP) 
#   input:
#       START
#       END
#       STEP
#   output:
#       sizes:  array of sizes of farms
def sizes_array (START,END,STEP) :                    
    sizes = []                            # List of sizes of farms
    for i in range (START,END,STEP):
        sizes.append (i)
    return sizes
Additional_sizes = sizes_array (START,END,STEP)      # solar farm sizes
CC_sizes = sizes_array (START,END,STEP)              # solar farm sizes
CT_sizes = sizes_array (START,END,STEP)              # solar farm sizes
RIC_sizes = sizes_array (START,END,STEP)             # solar farm sizes
ARO_sizes = sizes_array (START,END,STEP)             # solar farm sizes
Add_CC_sizes = sizes_array (START,END,STEP)          # solar farm sizes
Add_CT_sizes = sizes_array (START,END,STEP)          # solar farm sizes
Add_RIC_sizes = sizes_array (START,END,STEP)         # solar farm sizes
Add_ARO_sizes = sizes_array (START,END,STEP)         # solar farm sizes

# ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^

# ^^^^^^^^^^^^^ Energy output based on Size of ADDITIONAL resource ^^^^^^^^^^^^
# turn Add sizes list into matrix of sizes x 8760
# It gives a LIST (one row) & a MATRIX 
#   input:
#       sizes: array of sizes of farms
#       energyADD8760 : energy array 8760 for the additional renewable resource
#   output:
#       energy_list    : one row 8760 of the energy matrix
#       energy_matrix  : energy matrix sizes x energy_list
def energy_output (sizes, energyADD8760): 
    energy_matrix  = []
    for size in sizes:
        x= (size/100)    # x = multipier sizes/100 for solar (100MWh spreadsheet)
        energy_list=[]
        
        for hr in energyADD8760:
            y=x*hr
            energy_list.append(y)
        energy_matrix.append(energy_list)
    return energy_list, energy_matrix

# solar energy list & energy matrix produced based on solar farm sizes
Add_energy_l, Add_energy_m = energy_output (Additional_sizes, Additional8760)

CC_gas_energy_l, CC_gas_energy_m = energy_output (CC_sizes, CC_energyGas8760)
CT_gas_energy_l, CT_gas_energy_m = energy_output (CT_sizes, CT_energyGas8760)
RIC_gas_energy_l, RIC_gas_energy_m = energy_output (RIC_sizes, RIC_energyGas8760)
ARO_gas_energy_l, ARO_gas_energy_m = energy_output (ARO_sizes, ARO_energyGas8760)


# @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ gas @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
# ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
#                      adjust Matrix[size x 8760] with list [8760]
# ---------------adjust gas_energy_m with full_forecast8760----------
# --if there is enough MWh in row or matrix, to fill in row with --------
#                     full_forecast8760 then used 
# 
#   input:
#       energy_matrix: matrix[i,j] of arrays of energy based on sizes [i]
#       full_gap8760 : total energy gap needs to be filled in
#   output:
#       gas_energy_adj: the gas plan will not provided more MWh than the line 
#                       can take Matrix

def adjust_energy_gas (energy_Matrix, full_gap8760 ):
    gas_energy_adj = []          # this is a matrix [sizes x 8760]  
    for i in range(len(energy_Matrix)):    #  sizes rows
        adj=[]      # adjustment list, row of gas =[8760]
        
        # energy_Matrix[j]                  # sz = 8760 columnAdd
        for j in range(len(full_gap8760)):
            # Total Energy more than nameplate:
            if (energy_Matrix[i][j] >= full_gap8760[j]):
                adj.append(full_gap8760[j])
            else:
                 adj.append(energy_Matrix[i][j])
        gas_energy_adj.append(adj)
    return gas_energy_adj
CC_gas_energy_adj = adjust_energy_gas (CC_gas_energy_m, full_forecast8760 )
CT_gas_energy_adj = adjust_energy_gas (CT_gas_energy_m, full_forecast8760 )
RIC_gas_energy_adj = adjust_energy_gas (RIC_gas_energy_m, full_forecast8760 )
ARO_gas_energy_adj = adjust_energy_gas (ARO_gas_energy_m, full_forecast8760 )

# @@@@@@@@@@@@@@@@@@@@@@@@ end gas @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

# $$$$$$$$$$$$$$$$$$$ Existing $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$

# ========== Extra Energy accumulated from EXISTING resource (sum)============
# extra energy accumulated from existing resource, (energy8760[i] >= new_trans)
# list 8760 in [MWh]
#   input:
#       energyEXIST8760 : energy array 8760 for the EXISTING renewable resource
#       new_trans : energy needed based on new transmission (is conAddtant) 
#   output:
#       extra_existing  :sum extra energy of the existing resource, var
def extra_energy_sum (energyEXIST8760, new_trans ):
    extra_existing = 0           # initializing extra existing energy variable  
    for i in range(len(energyEXIST8760)):
        if (energyEXIST8760[i] >= new_trans):
            extra_existing += energyEXIST8760[i] - new_trans
    return extra_existing

# Existing sum of extra energy generated (variable)
extra_Exist = extra_energy_sum  (Existing8760, new_trans )

#$$$$$$$$$$$$$$$$$$$$$$$$$ end Existing $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$


# ///////////////////// curtail, extra, added, sizes x 8760 \\\\\\\\\\\\\\\\
# ^^^^^^^^ 1. Curtail matrix, 2. Extra Energy matrix, 3. Energy Added matrix  ^
#   1. Energy curtail ( full_forecast8760 < Additional8760 ), 
#   2. Energy_extra ( forecast8760 < Additional8760 < full_forecast8760) 
#   3. Energy added (Additional8760 < forecast8760) 
# Calculates the [MWh] output for a continuous of size farms
#   input:
#       full_forecast8760: 8760 list, gap from transmission line to existing resource 
#       forecast8760: 8760 list, gap from new_transmission line to existing resource
#       energy_matrix  : energy matrix sizes x energy_list
#   output:
#       curtail : curtail energy matrix [sizes x energy_list]
#       extra_additional : matrix [sizes x 8760] of sum of additional energy
#       added : added energy matrix [sizes x energy_list]

def energyCulcAdditionalResource (full_forecast8760, forecast8760, energy_matrix ):
   
    curtail =[]             # curtailment matrix
    extra_matrix= []        # extra energy matrix
    added = []              # Added energy matrix
    for i in range(len(energy_matrix)):    #  sizes rows
        curt=[]      # curtailment list, row of curtail =[]
        ad=[]        # solar energy added list, row of extra_matrix =[]
        extra = []   # extra energy added list, row of added =[]
        
        # energy_matrix[j]                  # sz = 8760 columnAdd
        for j in range(len(full_forecast8760)):
            # Total Energy more than nameplate:
            if (energy_matrix[i][j] > full_forecast8760[j]):
                curt.append(energy_matrix[i][j] - full_forecast8760[j])
                extra.append(full_forecast8760[j] - forecast8760[j])
                ad.append(forecast8760[j]) 
            #Total Energy between new_transmission & nameplate:
            elif (forecast8760[j] <= energy_matrix[i][j] <= full_forecast8760[j]):
                curt.append(0)
                extra.append(energy_matrix[i][j] - forecast8760[j])
                ad.append(forecast8760[j])
            #Total Energy smaller than new_transmission:  
            elif (energy_matrix[i][j] < forecast8760[j]):
                curt.append(0)
                extra.append(0)
                ad.append(energy_matrix[i][j])
        curtail.append(curt)
        extra_matrix.append(extra)
        added.append(ad)
            
    return curtail, extra_matrix, added

# cartail matrix, extra energy list & energy added matrix, by Additional resources 
Addition_curtail, extra_Add8760, Addition_added = \
    energyCulcAdditionalResource (full_forecast8760, forecast8760, Add_energy_m )

# gas cartail matrix, extra energy list & energy added matrix
CC_curtail, extra_CC8760, CC_added = \
    energyCulcAdditionalResource (full_forecast8760, forecast8760, CC_gas_energy_adj )
CT_curtail, extra_CT8760, CT_added = \
    energyCulcAdditionalResource (full_forecast8760, forecast8760, CT_gas_energy_adj )
RIC_curtail, extra_RIC8760, RIC_added = \
    energyCulcAdditionalResource (full_forecast8760, forecast8760, RIC_gas_energy_adj )
ARO_curtail, extra_ARO8760, ARO_added = \
    energyCulcAdditionalResource (full_forecast8760, forecast8760, ARO_gas_energy_adj )


# ///////////////////////////////////\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\
#                              Battery
 # ////// curtail, added, batt_storege_var, batt_use || list[sizes] & var \\\
#   1. Energy curtail ( full_forecast8760 < Additional8760 ), 
#   2. Energy added (Additional8760 < forecast8760)
#   
# Calculates the [MWh] output for a continuous of size farms + given size Battery
#      Energy_additional (list) + Energy Battery (list) is calculated separate 
#   input:
#       full_forecast8760: 8760 list, gap from transmission line to existing resource 
#       forecast8760: 8760 list, gap from new_transmission line to existing resource
#       energy_matrix  : energy matrix sizes x energy_list
#       size_battery: size of the battery will be added 
#   output: (turn things in to lists instead of matrices)
#       curtail_list : curtail list [sizes ]
#       added_list   : added energy list [sizes ]
#       charge_list: total stored energy by the battery unsued list [sizes ]
#       discharge_list :      total energy used by the battery list [sizes ]

bat_effic = .9    # efficiency of battery used
def energyAdditionalPlusBatteries \
    (full_forecast8760, forecast8760, energy_matrix, size_battery ): 
   
    bat_full = size_battery*4    # full battery charge (4hr bat*size_battery)
    # lists
    curtail_list =[]             # curtailment list [size]   
    added_list = []              # Added energy list [size] 
    charge_list = []             # battery charge list [size] 
    discharge_list = []          # battery discharge list [size] 
    for i in range(len(energy_matrix)):    #  sizes rows
        
        Sum_curt=0             # curtailment list, row of curtail =[]
        Sum_ad=0               # solar energy added list, row of extra_matrix =[]
        # battarie variebles collected at the end
        batt_storege_var  = 0  # initializing the total stored energy by the battery
        batt_use_var   = 0     # initializing the total battery used energy
        # energy_matrix[j]                  # sz = 8760 columnAdd
        for j in range(len(full_forecast8760)):     # 1) > the NP
            
            # 1. Energy >= than nameplate: ===== Charge phase ================
            if (energy_matrix[i][j] >= full_forecast8760[j]):     
                Sum_ad += full_forecast8760[j] # additional source energy sums
                
                # a) E_in > sz_bat ----------------------------------------------------------------------
                #    E_in = (energy_matrix[i][j]*bat_effic - full_forecast8760[j])
                if (energy_matrix[i][j]*bat_effic - full_forecast8760[j]) -size_battery >0:
                    
                    # if there is room avail. (room = bat_full- batt_storage)
                    if bat_full - batt_storege_var > 0:
                        # room > size_battery
                        if (bat_full - batt_storege_var) - size_battery > 0:
                            batt_storege_var += size_battery
                            Sum_curt += energy_matrix[i][j] - size_battery- full_forecast8760[j] 
                        # room < size_battery 
                        else: 
                            Sum_curt += energy_matrix[i][j] - (bat_full - batt_storege_var)- full_forecast8760[j] 
                            batt_storege_var += (bat_full - batt_storege_var)
                            
                    # if NO room avail        
                    else: 
                         Sum_curt += energy_matrix[i][j] - full_forecast8760[j] 
               # ------------------------------------------------------------------------------------------- 
                # b) E < sz_bat        
                else:
                    
                    # if there is room avail.  (room = bat_full- batt_storage)
                    if bat_full - batt_storege_var > 0:
                        # room available > Energy in, [Energy in = (energy_matrix[i][j]*bat_effic - full_forecast8760[j])]
                        if (bat_full - batt_storege_var) - (energy_matrix[i][j]*bat_effic - full_forecast8760[j]) > 0:
                            batt_storege_var += (energy_matrix[i][j]*bat_effic - full_forecast8760[j])
                            
                         # room available < Energy in
                        else:               # energy in                              -              room
                            Sum_curt += (energy_matrix[i][j] - full_forecast8760[j])  - (bat_full - batt_storege_var)
                            batt_storege_var +=(bat_full - batt_storege_var)
                            
                    # if NO room avail        
                    else:                        # energy in 
                         Sum_curt += energy_matrix[i][j] - full_forecast8760[j] 
                          
            # 2. Energy < than nameplate: =============== discharge phase ======
            else :
                Sum_ad += energy_matrix[i][j] 
                
                # a) f_for > sz_bat ----------------------------------------------------------------------
                if full_forecast8760[j] - size_battery > 0 :
                    
                    # if there is storage avail.
                    if batt_storege_var > 0:
                        # storage > size_battery
                        if batt_storege_var - size_battery >0:
                            batt_use_var += size_battery
                            batt_storege_var -= size_battery
                            
                        # storage < size_battery 
                        else: 
                            batt_use_var += batt_storege_var
                            batt_storege_var = 0
        
                    # if NO storage avail        
                    else: 
                         batt_storege_var = 0
               # ------------------------------------------------------------------------------------------- 
                # b) f_for < sz_bat ---------------------------------------------------------------------       
                else:
                    
                    # if storage avail.
                    if batt_storege_var > 0:
                        # storage > Energy need, (E_need = full_forecast8760[j])
                        if batt_storege_var - full_forecast8760[j] > 0 :
                            batt_use_var += full_forecast8760[j]
                            batt_storege_var -= full_forecast8760[j]
                            
                        # storage < Energy need
                        else:
                            batt_use_var += batt_storege_var
                            batt_storege_var = 0
                            
                    # if NO storage avail        
                    else: 
                         batt_storege_var = 0 
        
        # lists [size]        
        curtail_list.append(Sum_curt)
        added_list.append(Sum_ad)
        charge_list.append(batt_storege_var)
        discharge_list.append(batt_use_var)    
    return curtail_list, added_list, discharge_list, charge_list   

# Additional resource & battery informations
AddBat1_curtail, AddBat1_added ,batt1_use, batt1_storege = \
    energyAdditionalPlusBatteries (full_forecast8760, forecast8760, Add_energy_m, SZbat1 )
AddBat2_curtail, AddBat2_added ,batt2_use, batt2_storege = \
    energyAdditionalPlusBatteries (full_forecast8760, forecast8760, Add_energy_m, SZbat2 )
AddBat3_curtail, AddBat3_added ,batt3_use, batt3_storege = \
    energyAdditionalPlusBatteries (full_forecast8760, forecast8760, Add_energy_m, SZbat3 )

# =========================================================================
# ========== SUM Row [i] together in MATRIX, prodused list size [sizes] ==
# the sum of rows [i] 8760 in a matrix [size x 8760]
#   input:
#       matrix  : matrix [size x 8760]
#   output:
#       sum_row_list : i (j+j+j)...=total, list [sizes]

def matrixROWsummation (matrix):
    sum_row=0              # sum of each row of the matrix (var)
    row=[]                 # list of i rows in added matrix 
    sum_row_list = []      # list of sum column values [row] 
    for i in range(len( matrix)):
        row = matrix[i][:] # [8760] 
        sum_row = sum(float(j) for j in row)
        sum_row_list.append(sum_row)
    return sum_row_list
                         
# Total added below new_transmission levels [MWh] for varies sizes farms, LIST 8760
Sum_Add_under = matrixROWsummation (Addition_added) # list [sizes]

Sum_CC_under = matrixROWsummation (CC_added)   # list [sizes]
Sum_CT_under = matrixROWsummation (CT_added)   # list [sizes]
Sum_RIC_under = matrixROWsummation (RIC_added) # list [sizes]
Sum_ARO_under = matrixROWsummation (ARO_added) # list [sizes]


#                        TOTAL CURTAILMENT
# Total curail energy added [MWh] for varies sizes Additional farms, LIST[sizes] 
Sum_Addition_curtail = matrixROWsummation (Addition_curtail)


# extra energy (additional & gas ) TOTAL [MWh] for varies sizes farms
extra_Add = matrixROWsummation (extra_Add8760)        # list [sizes]

extra_CC = matrixROWsummation (extra_CC8760)          # list [sizes]
extra_CT = matrixROWsummation (extra_CT8760)          # list [sizes]
extra_RIC = matrixROWsummation (extra_RIC8760)        # list [sizes]
extra_ARO = matrixROWsummation (extra_ARO8760)        # list [sizes]

# ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^

# ========== SUM Function: list [sizes]+ variable ============================
# Adds a list sizes + variable
#   input:
#       list_sizes : list of size sizes
#       variable : value of 1 variable 
#   output:
#       listPLUSvar: list size [sizes] + variable
#Existing8760, Sum_Existing (var), extra_Exist  ()
def sum_list_var (list_sizes, variable):
    listPLUSvar =[]
    for i in list_sizes:
        listPLUSvar.append(variable + i)
    return  listPLUSvar

# Total extra energy additional for diff. sizes (list size) + Total extra Existing 
#                      Total extra | not used here |
extraTotAdd= sum_list_var (extra_Add,extra_Exist)

extraTotCC = sum_list_var (extra_CC,extra_Exist)
extraTotCT = sum_list_var (extra_CT,extra_Exist)
extraTotRIC = sum_list_var (extra_RIC,extra_Exist)
extraTotARO = sum_list_var (extra_ARO,extra_Exist)

#print ("\nextraTotAdd: ",extraTotAdd[nAdd1])
#print ("extraTotAdd): ",len(extraTotAdd))
# ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
# ========== SUM Function: listA [sizes]+ listB [sizes] =====================
# Adds a listA [sizes] + listB [sizes]
#   input:
#       listA : listA of size sizes
#       listB : listB of size sizes
#   output:
#       listPLUSlist: sum of 2 lists size [sizes] 

def sum_list_list (listA,listB ):
    sumOFlists =[]         
    x = 0
    for i in range(len(listA )):
        x= listA[i] + listB[i]
        sumOFlists.append(x)
    return  sumOFlists

# Tatal Energy of added [MWh] under + above utilization levels
# @@@@@@@@ 3 different sizes of Additional pass to M3 @@@@@@@@@@@@@@@@@@@@@@@@
Sum_Add_added = sum_list_list (Sum_Add_under,extra_Add )

# Tatal Energy of added [MWh] under + above utilization levels for gas
Sum_CC_added = sum_list_list (Sum_CC_under,extra_CC )
Sum_CT_added = sum_list_list (Sum_CT_under,extra_CT )
Sum_RIC_added = sum_list_list (Sum_RIC_under,extra_RIC )
Sum_ARO_added = sum_list_list (Sum_ARO_under,extra_ARO )

#                          battery
# Add together Enregy (charge + discharge) of battery List [size] + list [size]
Bat1_tot = sum_list_list (batt1_use, batt1_storege) 
Bat2_tot = sum_list_list (batt2_use, batt2_storege) 
Bat3_tot = sum_list_list (batt3_use, batt3_storege) 

# ---Total Energy Added-----  Additional (list) + Battery (use+ store) (Var) --
AddBat1_tot = sum_list_list (AddBat1_added, Bat1_tot )
AddBat2_tot = sum_list_list (AddBat2_added, Bat2_tot )
AddBat3_tot = sum_list_list (AddBat3_added, Bat3_tot )

# ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^

# ===========================================================================
# ========== test for percentage of line utilization =====================
# finds line utilization as a fraction
#   input:
#       Sum_existing : the yearly sum energy of the existing resource [MWh], var
#       Sum_additional_added : the yearly sum energy of the additonal [MWh], list[size]
#       Sum_nameplate : the yearly sum energy of the nameplate [MWh], var
#   output:
#       utilization_act: fraction of utilization list [sizes]
def actual_utilization (Sum_existing, Sum_additional_added, Sum_nameplate):
    utilization_act = []
    x = 0
    for i in range(len(Sum_additional_added)):
        x = ((Sum_existing + Sum_additional_added[i])/Sum_nameplate)
        utilization_act.append(x)
    return  utilization_act 
# fraction utilization ( produces a list [sizes] of utilazation basen on size)
utilization_Add = actual_utilization(Sum_Existing, Sum_Add_added, Sum_nameplate)
utilization_CC = actual_utilization(Sum_Existing, Sum_CC_added, Sum_nameplate)
utilization_CT = actual_utilization(Sum_Existing, Sum_CT_added, Sum_nameplate)
utilization_RIC = actual_utilization(Sum_Existing, Sum_RIC_added, Sum_nameplate)
utilization_ARO = actual_utilization(Sum_Existing, Sum_ARO_added, Sum_nameplate)

# Additional & Batterie Utilization
utilization_AddBat1 = actual_utilization(Sum_Existing, AddBat1_tot, Sum_nameplate)
utilization_AddBat2 = actual_utilization(Sum_Existing, AddBat2_tot, Sum_nameplate)
utilization_AddBat3 = actual_utilization(Sum_Existing, AddBat3_tot, Sum_nameplate)

# =========================================================================
# ===shifts by 1 to the Left & adds final value at last position again======
# ========= the comp_utilization created for computational purposes only =====
# creates a new_List, will be used to subtract original list (same size)
# removes 1st elemement, shifts 1 left, adds final value to last element
#   input:
#       List : original list
#   output:
#       new_List: will be used to subtract original list (same size)
def shift_one_add_final (List):
    x=List[-1]       # the last element placed again at end of list
    # it is shifted by 1, last element repeates, finds the y_i+1-y_i 
    new_List = List.copy()    
    new_List.append(x)
    del(new_List[0])           # keep size of list constant
    return new_List
comp_utilization_Add = shift_one_add_final (utilization_Add)
comp_utilization_CC = shift_one_add_final (utilization_CC)
comp_utilization_CT = shift_one_add_final (utilization_CT)
comp_utilization_RIC = shift_one_add_final (utilization_RIC)
comp_utilization_ARO = shift_one_add_final (utilization_ARO)
# batteries
comp_utilization_AddBat1 = shift_one_add_final (utilization_AddBat1)
comp_utilization_AddBat2 = shift_one_add_final (utilization_AddBat2)
comp_utilization_AddBat3 = shift_one_add_final (utilization_AddBat3)

# ==========================================================================
#                           Derivative
# finds the slope derivative of a list
# used for 1st & 2nd derivative
#   input:
#       STEP : Step size
#       List: original list
#       new_List: list used to subtract original list
#   output:
#       derivat: derivative list size 
def derivative (STEP, List, new_List ):   
    derivat = []             # derivative list
    y = 0
   
    for i in range(len(List)):
        y= (new_List[i]-List[i])/STEP
        derivat.append(y)
        
    return derivat

Additional_FstDeriv = derivative (STEP, utilization_Add,comp_utilization_Add )
Add_comp_FstDeriv= shift_one_add_final (Additional_FstDeriv )
Additional_SecDeriv= derivative (STEP, Additional_FstDeriv ,Add_comp_FstDeriv )

CC_FstDeriv = derivative (STEP, utilization_CC,comp_utilization_CC )
CC_comp_FstDeriv= shift_one_add_final (CC_FstDeriv )
CC_SecDeriv= derivative (STEP, CC_FstDeriv ,CC_comp_FstDeriv )

CT_FstDeriv = derivative (STEP, utilization_CT,comp_utilization_CT )
CT_comp_FstDeriv= shift_one_add_final (CT_FstDeriv )
CT_SecDeriv= derivative (STEP, CT_FstDeriv ,CT_comp_FstDeriv )

RIC_FstDeriv = derivative (STEP, utilization_RIC,comp_utilization_RIC )
RIC_comp_FstDeriv= shift_one_add_final (RIC_FstDeriv )
RIC_SecDeriv= derivative (STEP, RIC_FstDeriv ,RIC_comp_FstDeriv )

ARO_FstDeriv = derivative (STEP, utilization_ARO,comp_utilization_ARO )
ARO_comp_FstDeriv= shift_one_add_final (ARO_FstDeriv )
ARO_SecDeriv= derivative (STEP, ARO_FstDeriv ,ARO_comp_FstDeriv )

# batteries 1
AddBat1_FstDeriv = derivative (STEP, utilization_AddBat1,comp_utilization_AddBat1 )
AddBat1_comp_FstDeriv= shift_one_add_final (AddBat1_FstDeriv )
AddBat1_SecDeriv= derivative (STEP, AddBat1_FstDeriv ,AddBat1_comp_FstDeriv )
# batteries 2
AddBat2_FstDeriv = derivative (STEP, utilization_AddBat2,comp_utilization_AddBat2 )
AddBat2_comp_FstDeriv= shift_one_add_final (AddBat2_FstDeriv )
AddBat2_SecDeriv= derivative (STEP, AddBat2_FstDeriv ,AddBat2_comp_FstDeriv )
# batteries 3
AddBat3_FstDeriv = derivative (STEP, utilization_AddBat3,comp_utilization_AddBat3 )
AddBat3_comp_FstDeriv= shift_one_add_final (AddBat3_FstDeriv )
AddBat3_SecDeriv= derivative (STEP, AddBat3_FstDeriv ,AddBat3_comp_FstDeriv )


# ---Test = only looks in to 2nd derivative within the appropriate renange -
# ----------------- when END = 2 * nameplate ------------------------
Test = int(.9 * N) 

# -----for best choice only examines from 0 to .9 * N (total number of steps)---
# ------------------- when END = 2 * nameplate -------------------------------
# ---------- the rest of the 2nd derivitate is filled with 0 -----------------
# ----------------- prvents calulational errors -----------------------------
#   input:
#       Sec_derivat: the list [...] with 2nd derivatives
#   output:
#       add_zeros: last 2 elements in list [sizes] position with [..0,0]

def adjSecDeriv (Sec_derivat, Test):  
    zero_list = Sec_derivat[Test:]
    Sec_derivat = Sec_derivat[0:Test]
    #print ("len(Sec_derivat): ", len (Sec_derivat))
    #print ('len(zero_list): ', len(zero_list))
    for i in range(len(zero_list)):
        Sec_derivat.append(0)
         
    return Sec_derivat

Additional_SecDeriv= adjSecDeriv (Additional_SecDeriv, Test)

CC_SecDeriv= adjSecDeriv (CC_SecDeriv, Test)
CT_SecDeriv= adjSecDeriv (CT_SecDeriv, Test)
RIC_SecDeriv= adjSecDeriv (RIC_SecDeriv, Test)
ARO_SecDeriv= adjSecDeriv (ARO_SecDeriv, Test)
# batteries
AddBat1_SecDeriv = adjSecDeriv (AddBat1_SecDeriv, Test)
AddBat2_SecDeriv = adjSecDeriv (AddBat2_SecDeriv, Test)
AddBat3_SecDeriv = adjSecDeriv (AddBat3_SecDeriv, Test)
#print ("\nAdditional_SecDeriv: ", round (Additional_SecDeriv[nAdd1],8))
#print ("len(Additional_SecDeriv): ", len (Additional_SecDeriv))


# ===========================================================================
# === Suggests best choice additional resource based on 2nd derivative ======
#             * only use when 'var' at GUI is chosen *
#   input:
#       START: starting size of farm used in sizes array
#       STEP : step oe increasing farm sizes
#       derivat: list of 2nd derivative [sizes] 
#       var: variable that when chosen to display ideal it does if not 0
#       minus_add_comp: minus the position of the addtion resource if present
#   output:
#       choice :  suggested farm size  (var)
#       index:    position of farm size in size list (var)
#       sz:       size of the farm recommended (var)

def best_choice (START,STEP, derivat, var, minus_add_comp):
    choice = 0 
    n = 0
    for i in range(len( derivat)):
        
        if ( -derivat[i]  > choice):
            choice = -derivat [i] * var
            n = (i* var) - minus_add_comp
            
            # if the n is (-) d/t - minus_add_comp then n= 0
            if n <=0: 
                n = 0
            
    sz = n * STEP + START   

    return choice, n, sz

# additional ideal gas
Additional_choice, nAdd, Add_sz =  best_choice \
    (START,STEP, Additional_SecDeriv, 1, 0 ) 
# gas ideal gas    
CC_choice, nCC, CC_sz = best_choice (START,STEP, CC_SecDeriv, CC_Id, 0)  
CT_choice, nCT, CT_sz = best_choice (START,STEP, CT_SecDeriv, CT_Id, 0 )  
RIC_choice, nRIC, RIC_sz = best_choice (START,STEP, RIC_SecDeriv, RIC_Id, 0 )  
ARO_choice, nARO, ARO_sz = best_choice (START,STEP, ARO_SecDeriv, ARO_Id, 0 )  
# battery  
AddBat1_choice, nAddBat1, AddBat1_sz = best_choice\
    (START,STEP, AddBat1_SecDeriv, Bat1_Id, 0)  
AddBat2_choice, nAddBat2, AddBat2_sz = best_choice \
    (START,STEP, AddBat2_SecDeriv, Bat2_Id, 0)  
AddBat3_choice, nAddBat3, AddBat3_sz = best_choice \
    (START,STEP, AddBat3_SecDeriv, Bat3_Id, 0)  

# @@@@@@@@@@@@@@@@@@@@ monthly gas energy output @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
# @@@@@@@@@@@@@@@@@@@@ single gas input to M3    @@@@@@@@@@@@@@@@@@@@@@@@@@@@@

# hrs attributed to each month in the 8760 yearly period:
jen = 31*24; feb = jen+28*24;  mar= feb+31*24; april = mar+30*24
may= april+31*24; june= may+30*24; jul = june+31*24; aug=jul+31*24
sept=aug+30*24; octb= sept+31*24; nov= octb+30*24; dec= nov+31*24 


# determines the total Energy produced each month  
# input:
#   gas_energy_adj: the energy of gas as adjusted to full forcast and MP
# output :
#  Total energy output per month [MWh], var, based on generator size
def monthEnergyGas (gas_energy_adj,n):
    Jen = 0; Feb = 0;  Mar = 0; April = 0; May = 0; June = 0;
    Jul = 0; Aug = 0; Sept = 0; Octb = 0; Nov = 0; Dec = 0; 
    for hr in range(len(gas_energy_adj[n])):
        if hr < jen:
            Jen += gas_energy_adj[n][hr]
        if jen <= hr < feb:
            Feb += gas_energy_adj[n][hr]
        if feb <= hr <  mar:
            Mar += gas_energy_adj[n][hr]
        if mar <= hr <  april:
            April += gas_energy_adj[n][hr]
        if april <= hr <  may:
            May += gas_energy_adj[n][hr]
        if may <= hr <  june:
            June += gas_energy_adj[n][hr]
        if june <= hr <  jul:
            Jul += gas_energy_adj[n][hr]
        if jul <= hr <  aug:
            Aug += gas_energy_adj[n][hr]
        if aug <= hr <  sept:
            Sept += gas_energy_adj[n][hr]
        if sept <= hr <  octb:
            Octb += gas_energy_adj[n][hr]
        if octb <= hr <  nov:
            Nov += gas_energy_adj[n][hr]
        if nov <= hr <  dec:
            Dec += gas_energy_adj[n][hr]
    return Jen, Feb, Mar, April, May, June, Jul, Aug, Sept, Octb, Nov, Dec

# ------------------------3 single gas choices passed to M3 ------------------
# ------------ for ideal size (code recomendation) ---------------------------
JrCC, FbCC, MrCC, ApCC, MyCC, JnCC, JlCC, AgCC, SpCC, OcCC, NvCC, \
    DcCC = monthEnergyGas (CC_gas_energy_adj,nCC) 
JrCT, FbCT, MrCT, ApCT, MyCT, JnCT, JlCT, AgCT, SpCT, OcCT, NvCT, \
    DcCT = monthEnergyGas (CT_gas_energy_adj,nCT) 
JrRIC, FbRIC, MrRIC, ApRIC, MyRIC, JnRIC, JlRIC, AgRIC, SpRIC, OcRIC, NvRIC, \
    DcRIC = monthEnergyGas (RIC_gas_energy_adj,nRIC) 
JrA, FbA, MrA, ApA, MyA, JnA, JlA, AgA, SpA, OcA, NvA, \
    DcA = monthEnergyGas (ARO_gas_energy_adj,nARO) 

# ---------- for user choice 1 size -----------------------------------------     
JrCC1, FbCC1, MrCC1, ApCC1, MyCC1, JnCC1, JlCC1, AgCC1, SpCC1, OcCC1, NvCC1, \
    DcCC1 = monthEnergyGas (CC_gas_energy_adj,nCC1) 
JrCT1, FbCT1, MrCT1, ApCT1, MyCT1, JnCT1, JlCT1, AgCT1, SpCT1, OcCT1, NvCT1, \
    DcCT1 = monthEnergyGas (CT_gas_energy_adj,nCT1) 
JrRIC1, FbRIC1, MrRIC1, ApRIC1, MyRIC1, JnRIC1, JlRIC1, AgRIC1, SpRIC1, OcRIC1, NvRIC1, \
    DcRIC1 = monthEnergyGas (RIC_gas_energy_adj,nRIC1) 
JrA1, FbA1, MrA1, ApA1, MyA1, JnA1, JlA1, AgA1, SpA1, OcA1, NvA1, \
    DcA1 = monthEnergyGas (ARO_gas_energy_adj,nARO1) 

# ------- for user choice 2 size --------------------------------------------     
JrCC2, FbCC2, MrCC2, ApCC2, MyCC2, JnCC2, JlCC2, AgCC2, SpCC2, OcCC2, NvCC2, \
    DcCC2 = monthEnergyGas (CC_gas_energy_adj,nCC2) 
JrCT2, FbCT2, MrCT2, ApCT2, MyCT2, JnCT2, JlCT2, AgCT2, SpCT2, OcCT2, NvCT2, \
    DcCT2 = monthEnergyGas (CT_gas_energy_adj,nCT2) 
JrRIC2, FbRIC2, MrRIC2, ApRIC2, MyRIC2, JnRIC2, JlRIC2, AgRIC2, SpRIC2, OcRIC2, NvRIC2, \
    DcRIC2 = monthEnergyGas (RIC_gas_energy_adj,nRIC2) 
JrA2, FbA2, MrA2, ApA2, MyA2, JnA2, JlA2, AgA2, SpA2, OcA2, NvA2, \
    DcA2 = monthEnergyGas (ARO_gas_energy_adj,nARO2) 

# ^^^^^^^^^^^^^^^^^^^^^^end^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^

# @@@@@@@@@@@@@@@@@@@@@@@@@@@ start Additional + gas @@@@@@@@@@@@@@@@@@@@@@@@@@

# &&&&&&&&&&&&&&&&&&&              key            $$$$$$$$$$$$$$$$$$$$$$$$$$$$
# ideal    = Add_ideal  + gas_ideal
# choice 1 = Add_choice + gas_ideal
# choice 2 = Add_ideal  + gas_choice
# choice 3 = Add_choice + gas_choice
# $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$

# use Add_sz to find the energy output of the best size Additional resource

# &&&&&&&&&&&&&& using Ideal added and in order to find desired gas &&&&&&&&&&&

# ******************** Ideal Additional ***********************************
# slice [8760] of ideal Additional Energy [8760] produced based on profile
# ^^^^^^^^^^^^^ Energy output based on Size of ADDITIONAL resource ^^^
# It gives a LIST [8760] of the ideal size additional resource 8760 ptofile
#   input:
#       sizes: var of farm size
#       energyADD8760 : energy array 8760 for the additional renewable resource
#   output:
#       ideal_energy    : one row 8760 of the energy matrix
def ideal_energy_output (sizes, energyADD8760): 
    # initializing list and var
    ideal_energy_list  = []
    y=0                
    for hr in energyADD8760:
        y=(sizes/100)*hr      # sizes/100 for solar (100MWh spreadsheet)
        ideal_energy_list.append(y)
    return ideal_energy_list

# solar energy list[8760], when ideal size farm is used in [MWh]
idealAdd_energy = ideal_energy_output ( Add_sz, Additional8760)
cho1Add_energy = ideal_energy_output ( SZadd1, Additional8760)
  

# ///////////////////// list 8760 curtail, extra, added \\\\\\\\\\\\\\\\\\\\\
# =====for ideal Additional --- find (curtail[8760], extra[8760], added[8760]) 
#   1. Energy curtail ( full_forecast8760 < Additional8760 ), 
#   2. extra ( forecast8760 < Additional8760 < full_forecast8760) 
#   3. added (Additional8760 < forecast8760) 
# Examines the Existing + ideal Additional gives list 8760
#   input:
#       ideal_energy_list: 8760 list, ideal Additional 
#       full_forecast8760: 8760 list, gap of nameplate - Existing 
#       forecast8760  : 8760 list, gap after factor * nameplate
#   output:
#       curtail : curtail energy list [8760]
#       extra : extra list [8760] between full_forecast8760 & forecast8760
#       added : list [8760] of added energy below forecast8760

def IdealCurtExtraAdd (ideal_energy_list, full_forecast8760, forecast8760 ):
    cartail = []  # curtail list 8760 
    extra = []    # extra list [8760] between full_forecast8760 & forecast8760
    added = []    # list [8760] of added energy below forecast8760
    for i in range(len(ideal_energy_list)):
        if (ideal_energy_list[i] > full_forecast8760[i] ):
            cartail.append(ideal_energy_list[i] - full_forecast8760[i])
            extra.append(full_forecast8760[i]-forecast8760[i])
            added.append(forecast8760[i])
        elif (full_forecast8760[i] >=ideal_energy_list[i] >= forecast8760[i] ):
            cartail.append(0)
            extra.append(ideal_energy_list[i] - full_forecast8760[i] )
            added.append(forecast8760[i])
        elif (forecast8760[i] >= ideal_energy_list[i] ):
            cartail.append(0)
            extra.append(0)
            added.append(forecast8760[i] - ideal_energy_list[i])
    return cartail, extra, added


# ideal Additional [8760] and Existing [8760] find:
#   curtail, extra & added of the Additional resource
ideal_Add_curtail8760, ideal_Add_extra8760, ideal_Add_added8760 = \
    IdealCurtExtraAdd (idealAdd_energy, full_forecast8760, forecast8760 )
# choice 1
cho1_Add_curtail8760, cho1_Add_extra8760, cho1_Add_added8760 = \
    IdealCurtExtraAdd (cho1Add_energy, full_forecast8760, forecast8760 )
    

# ================== list [8760] to sum var =================================
# ========== Energy curtail, extra, added from ideal additional============
# extra energy accumulated from existing resource, (energy8760[i] >= new_trans)
# list 8760 in [MWh]
#   input:
#       listA : list 8760 to be sum in to a var
#   output:
#      var  :sum of list[8760]
def SumListToVar (listA):
    var = 0           # initializing extra existing energy variable  
    for i in range(len(listA)):
        var += listA[i] 
    return var

# Ideal Additional sum of extra energy generated (variable)
Sum_ideal_curtail = SumListToVar (ideal_Add_curtail8760) # var = Total curt
Sum_ideal_extra =  SumListToVar (ideal_Add_extra8760)    # var
Sum_ideal_added = SumListToVar (ideal_Add_added8760)     # var
# choice 1
Sum_cho1_curtail = SumListToVar (cho1_Add_curtail8760) # var = Total curt
Sum_cho1_extra =  SumListToVar (cho1_Add_extra8760)    # var
Sum_cho1_added = SumListToVar (cho1_Add_added8760)     # var

# *****************************************************************************

# $$$$$$$$$$$$$$$$$$$$$ combinational gas $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
# ********************* List + List ***************************************
#         List + List (Existing list [8760] + tested Additional list [8760] )
#                   = SUM_Exist_Add (list of 8760) 
# Existing + best Additional = new energy list 8760
SUM_Exist_Add = sum_list_list (Existing8760, idealAdd_energy ) # ideal
SUM_Exist_Add1 = sum_list_list (Existing8760, cho1Add_energy ) # choice 1
#print ("\nSUM_Exist_Add : \n", SUM_Exist_Add [a:b])

# ///////////////////// list 8760 curtail, full, forcast \\\\\\\\\\\\\\\\\\\\\
# = when Existing + Added --- find (curtail 8760, new full_forcast, new_forcast) ==
#   1. Energy curtail ( full_forecast8760 < Additional8760 ), 
#   2. full gap ( forecast8760 < Additional8760 < full_forecast8760) 
#   3. new gap (Additional8760 < forecast8760) 
# Examines the best Existing + Additional gives list 8760
# fn to find cartail and gap needed to be filled by gas (sizes x 8760)
#   input:
#       SUM_Exist_Add: 8760 list, Existing + Additional 
#       nameplate8760: 8760 list, of constant nameplate
#       new_trans8760  : 8760 list, constant utilization factor * nameplate
#   output:
#       curtail : curtail energy list [8760]
#       new_full_forcast : list [8760] of gap needed to be filled
#       new_forcast : list [8760] of gap needed to be filled below new transm
def newCurtForcVal (SUM_Exist_Add, nameplate8760, new_trans8760 ):
    cartail = []          # curtail list 8760 (the curtail should be the same as ideal_Add_curtail8760)
    new_full_forcast = [] # new_full_forcast 8760 (gap needs to filled by gas)
    new_forcast = []      # new_forcast 8760 (gap needs to filled by gas)
    for i in range(len(nameplate8760)):
        if (SUM_Exist_Add[i] > nameplate8760[i] ):
            cartail.append(SUM_Exist_Add[i] - nameplate8760[i])
            new_full_forcast.append(0)
            new_forcast.append(0)
        elif (nameplate8760[i] >= SUM_Exist_Add[i] >= new_trans8760[i] ):
            cartail.append(0)
            new_full_forcast.append(nameplate8760[i] - SUM_Exist_Add[i] )
            new_forcast.append(0)
        elif (new_trans8760[i] >= SUM_Exist_Add[i] ):
            cartail.append(0)
            new_full_forcast.append(nameplate8760[i] - SUM_Exist_Add[i] )
            new_forcast.append(new_trans8760[i] - SUM_Exist_Add[i] )
    return cartail, new_full_forcast, new_forcast

# the Additional (wind or solar) is the only curtail here
# ideal
Exist_Add_curtail8760, Exist_Add_new_full_forc8760, Exist_Add_new_forc8760 = \
    newCurtForcVal (SUM_Exist_Add, nameplate8760, new_trans8760 )
# choice 1    
Exist_Add1_curtail8760, Exist_Add1_new_full_forc8760, Exist_Add1_new_forc8760 = \
    newCurtForcVal (SUM_Exist_Add1, nameplate8760, new_trans8760 )

# ^^^^^^^^^^^ adjust Matrix[size x 8760] with list [8760] ^^^^^^^^^^^^^^^^^^^
# ---------------adjust gas_energy_m with Exist_Add_new_full_forc8760 -----
# --if there is enough MWh in row or matrix, to fill in row with --------
#                     Exist_Add_new_full_forc8760 then used 
# 
#   input:
#       energy_matrix: matrix[i,j] of arrays of energy based on sizes [i]
#       Exist_Add_new_full_forc8760 : total energy gap needs to be filled in
#   output:
#       gas_energy_adj: the gas plan will not provided more MWh than the line 
#                       can take
# ideal
AddCC_gas_energy_adj = adjust_energy_gas (CC_gas_energy_m, Exist_Add_new_full_forc8760)
AddCT_gas_energy_adj = adjust_energy_gas (CT_gas_energy_m, Exist_Add_new_full_forc8760 )
AddRIC_gas_energy_adj = adjust_energy_gas (RIC_gas_energy_m, Exist_Add_new_full_forc8760)
AddARO_gas_energy_adj = adjust_energy_gas (ARO_gas_energy_m, Exist_Add_new_full_forc8760)
# choice 1 
AddCC1_gas_energy_adj = adjust_energy_gas (CC_gas_energy_m, Exist_Add1_new_full_forc8760)
AddCT1_gas_energy_adj = adjust_energy_gas (CT_gas_energy_m, Exist_Add1_new_full_forc8760 )
AddRIC1_gas_energy_adj = adjust_energy_gas (RIC_gas_energy_m, Exist_Add1_new_full_forc8760)
AddARO1_gas_energy_adj = adjust_energy_gas (ARO_gas_energy_m, Exist_Add1_new_full_forc8760)

# //////////////////// Main fn that give matrix [size x 8760] \\\\\\\\\\\\\\
#                     curtail, extra, added matrix size x 8760 

# ^^ 1. Curtail matrix = 0, 2. Extra Energy matrix, 3. Energy Added matrix  ^^^
#   1. Energy curtail ( full_forecast8760 < Additional8760 ), 
#   2. Energy_extra ( forecast8760 < Additional8760 < full_forecast8760) 
#   3. Energy added (Additional8760 < forecast8760) 
# Calculates the [MWh] output for a continuous of size farms
#   input:
#       Exist_Add_new_full_forc8760: 8760 list, gap transmission line -ExistAdd 
#       Exist_Add_new_forc8760: 8760 list, gap from new_transmission line -ExistAdd 
#       gas_energy_adj: energy matrix sizes x 8760
#   output:
#       curtail : curtail energy matrix [sizes x energy_list] - here should be 0
#       extra_additional : matrix [sizes x 8760] of sum of additional energy
#       added : added energy matrix [sizes x energy_list]
# ideal
Add_CC_curtail, extra_Add_CC8760, Add_CC_added = \
    energyCulcAdditionalResource \
        (Exist_Add_new_full_forc8760, Exist_Add_new_forc8760, AddCC_gas_energy_adj )
Add_CT_curtail, extra_Add_CT8760, Add_CT_added = \
    energyCulcAdditionalResource \
        (Exist_Add_new_full_forc8760, Exist_Add_new_forc8760, AddCT_gas_energy_adj )
Add_RIC_curtail, extra_Add_RIC8760, Add_RIC_added = \
    energyCulcAdditionalResource \
        (Exist_Add_new_full_forc8760, Exist_Add_new_forc8760, AddRIC_gas_energy_adj )
Add_ARO_curtail, extra_Add_ARO8760, Add_ARO_added = \
    energyCulcAdditionalResource \
        (Exist_Add_new_full_forc8760, Exist_Add_new_forc8760, AddARO_gas_energy_adj )
# choice 1 
Add1_CC_curtail, extra_Add1_CC8760, Add1_CC_added = \
    energyCulcAdditionalResource \
        (Exist_Add1_new_full_forc8760, Exist_Add1_new_forc8760, AddCC1_gas_energy_adj )
Add1_CT_curtail, extra_Add1_CT8760, Add1_CT_added = \
    energyCulcAdditionalResource \
        (Exist_Add1_new_full_forc8760, Exist_Add1_new_forc8760, AddCT1_gas_energy_adj )
Add1_RIC_curtail, extra_Add1_RIC8760, Add1_RIC_added = \
    energyCulcAdditionalResource \
        (Exist_Add1_new_full_forc8760, Exist_Add1_new_forc8760, AddRIC1_gas_energy_adj )
Add1_ARO_curtail, extra_Add1_ARO8760, Add1_ARO_added = \
    energyCulcAdditionalResource \
        (Exist_Add1_new_full_forc8760, Exist_Add1_new_forc8760, AddARO1_gas_energy_adj )


# =========================================================================
# ========== SUM Row [i] together in MATRIX, prodused list size [sizes] ==
# the sum of rows [i] 8760 in a matrix [size x 8760]
#   input:
#       matrix  : matrix [size x 8760]
#   output:
#       sum_row_list : i (j+j+j)...=total, list [sizes]

# Total 'added below new_trans levels' [MWh] for diff sizes gas farms, LIST[sizes]
# ideal
Sum_Add_CC_under = matrixROWsummation (Add_CC_added)   # list [sizes]
Sum_Add_CT_under = matrixROWsummation (Add_CT_added)   # list [sizes]
Sum_Add_RIC_under = matrixROWsummation (Add_RIC_added) # list [sizes]
Sum_Add_ARO_under = matrixROWsummation (Add_ARO_added) # list [sizes]
# choice 1 
Sum_Add1_CC_under = matrixROWsummation (Add1_CC_added)   # list [sizes]
Sum_Add1_CT_under = matrixROWsummation (Add1_CT_added)   # list [sizes]
Sum_Add1_RIC_under = matrixROWsummation (Add1_RIC_added) # list [sizes]
Sum_Add1_ARO_under = matrixROWsummation (Add1_ARO_added) # list [sizes]

#print ("\nSum_Add_CC_under): ", Sum_Add_CC_under[nAddCC1])
#print ("len(Sum_Add_CC_under): ", len(Sum_Add_CC_under))


# 'Total curail energy added [MWh]' for varies sizes gas farms, LIST[sizes]
# it must be 0, IGNORE IT
Sum_Add_CC_curtail = matrixROWsummation (Add_CC_curtail)    # ideal
Sum_Add1_CC_curtail = matrixROWsummation (Add1_CC_curtail)  # choice 1


# 'extra energy gas TOTAL [MWh]' for varies sizes gas farms
# ideal
Sum_extra_Add_CC = matrixROWsummation (extra_Add_CC8760)       # list [sizes]
Sum_extra_Add_CT = matrixROWsummation (extra_Add_CT8760)       # list [sizes]
Sum_extra_Add_RIC = matrixROWsummation (extra_Add_RIC8760)     # list [sizes]
Sum_extra_Add_ARO = matrixROWsummation (extra_Add_ARO8760)     # list [sizes]
# choice 1
Sum_extra_Add1_CC = matrixROWsummation (extra_Add1_CC8760)     # list [sizes]
Sum_extra_Add1_CT = matrixROWsummation (extra_Add1_CT8760)     # list [sizes]
Sum_extra_Add1_RIC = matrixROWsummation (extra_Add1_RIC8760)   # list [sizes]
Sum_extra_Add1_ARO = matrixROWsummation (extra_Add1_ARO8760)   # list [sizes]

#print ("\nSum_extra_Add_CC: ", round(Sum_extra_Add_CC[nAddCC1],1))
#print ("len(Sum_extra_Add_CC): ", len(Sum_extra_Add_CC))

# ========== SUM Function: listA [sizes]+ listB [sizes] =====================
# Adds a listA [sizes] + listB [sizes]
#   input:
#       listA : listA of size sizes
#       listB : listB of size sizes
#   output:
#       listPLUSlist: sum of 2 lists size [sizes] 

# Tatal Energy of added [MWh] 'under + above' utilization levels for 'gas'
# ideal
Sum_Add_CC_added = sum_list_list (Sum_Add_CC_under, Sum_extra_Add_CC )
Sum_Add_CT_added = sum_list_list (Sum_Add_CT_under, Sum_extra_Add_CT)
Sum_Add_RIC_added = sum_list_list (Sum_Add_RIC_under, Sum_extra_Add_RIC )
Sum_Add_ARO_added = sum_list_list (Sum_Add_ARO_under, Sum_extra_Add_ARO )
# choice 1
Sum_Add1_CC_added = sum_list_list (Sum_Add1_CC_under, Sum_extra_Add1_CC )
Sum_Add1_CT_added = sum_list_list (Sum_Add1_CT_under, Sum_extra_Add1_CT)
Sum_Add1_RIC_added = sum_list_list (Sum_Add1_RIC_under, Sum_extra_Add1_RIC )
Sum_Add1_ARO_added = sum_list_list (Sum_Add1_ARO_under, Sum_extra_Add1_ARO )


# $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
# ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
#                       | Not used here |
# ========== SUM Function: list [sizes]+ variable ============================
# Adds a list sizes + variable
#   input:
#       list_sizes : list of size sizes
#       variable : value of 1 variable 
#   output:
#       listPLUSvar: list size [sizes] + variable
# Extra gas list [sizes] + Extra (Existing + ideal added) var

# Total extra energy gas diff. sizes [sizes] + Total extra Existing + ideal Add 
#                      Total extra | not used here |
extraTotExistAddition = extra_Exist + Sum_ideal_extra # var,ideal
extraTotExistAddition1 = extra_Exist + Sum_cho1_extra # var, choice 1

# ideal 'Total extra in [MWh]'
extraTotCC = sum_list_var (Sum_extra_Add_CC,extraTotExistAddition)
extraTotCT = sum_list_var (Sum_extra_Add_CT,extraTotExistAddition)
extraTotRIC = sum_list_var (Sum_extra_Add_RIC,extraTotExistAddition)
extraTotARO = sum_list_var (Sum_extra_Add_ARO,extraTotExistAddition)
# choice 1 'Total extra in [MWh]'
extraTotCC1 = sum_list_var (Sum_extra_Add1_CC,extraTotExistAddition1)
extraTotCT1 = sum_list_var (Sum_extra_Add1_CT,extraTotExistAddition1)
extraTotRIC1 = sum_list_var (Sum_extra_Add1_RIC,extraTotExistAddition1)
extraTotARO1 = sum_list_var (Sum_extra_Add1_ARO,extraTotExistAddition1)

# ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^

# !!!!!!!!!!!!!!!!!!  Combined (Additioanl + gas) !!!!!!!!!!!!!!!!!!!!!!!!!!!
# ========== SUM Function: list [sizes]+ variable ============================
# Adds a list sizes + variable
#   input:
#       list_sizes : list of size sizes
#       variable : value of 1 variable 
#   output:
#       listPLUSvar: list size [sizes] + variable
# Extra gas list [sizes] + Extra (Existing + ideal added) var

# Total energy of additional resource Added (added + extra) = var
Sum_Additional_added = Sum_ideal_added + Sum_ideal_extra # var, ideal
Sum_Additional1_added = Sum_cho1_added + Sum_cho1_extra  # var, choice 1

# Total energy added (Addditional + gas) (added + extra) = list[sizes] + var
#ideal
Sum_Add_CC_added = sum_list_var (Sum_Add_CC_added, Sum_Additional_added )
Sum_Add_CT_added = sum_list_var (Sum_Add_CT_added, Sum_Additional_added )
Sum_Add_RIC_added = sum_list_var (Sum_Add_RIC_added, Sum_Additional_added )
Sum_Add_ARO_added = sum_list_var (Sum_Add_ARO_added, Sum_Additional_added )
# choice 1
Sum_Add1_CC_added = sum_list_var (Sum_Add1_CC_added, Sum_Additional1_added )
Sum_Add1_CT_added = sum_list_var (Sum_Add1_CT_added, Sum_Additional1_added )
Sum_Add1_RIC_added = sum_list_var (Sum_Add1_RIC_added, Sum_Additional1_added )
Sum_Add1_ARO_added = sum_list_var (Sum_Add1_ARO_added, Sum_Additional1_added )

# !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

# &&&&&&&&&&& Utilization & Best Choice (Existing + Additional + gas) &&&&&&&&
# ===========================================================================
# ========== test for percentage of line utilization =====================
# finds line utilization as a fraction
#   input:
#       Sum_existing : the yearly sum energy of the existing resource [MWh], var
#       Sum_additional_added : the yearly sum energy of the additonal [MWh], list[size]
#       Sum_nameplate : the yearly sum energy of the nameplate [MWh], var
#   output:
#       utilization_act: fraction of utilization list [sizes]

# fraction utilization ( produces a list [sizes] of utilazation based on size)
# (Existing + ((Additional + gas)[size])) / Sum_nameplate
# ideal
utilization_Add_CC = actual_utilization(Sum_Existing, Sum_Add_CC_added, Sum_nameplate)
utilization_Add_CT = actual_utilization(Sum_Existing, Sum_Add_CT_added, Sum_nameplate)
utilization_Add_RIC = actual_utilization(Sum_Existing, Sum_Add_RIC_added, Sum_nameplate)
utilization_Add_ARO = actual_utilization(Sum_Existing, Sum_Add_ARO_added, Sum_nameplate)
# choice 1
utilization_Add1_CC = actual_utilization(Sum_Existing, Sum_Add1_CC_added, Sum_nameplate)
utilization_Add1_CT = actual_utilization(Sum_Existing, Sum_Add1_CT_added, Sum_nameplate)
utilization_Add1_RIC = actual_utilization(Sum_Existing, Sum_Add1_RIC_added, Sum_nameplate)
utilization_Add1_ARO = actual_utilization(Sum_Existing, Sum_Add1_ARO_added, Sum_nameplate)


# =========================================================================
# ===shifts by 1 to the Left & adds final value at last position again======
# ========= the comp_utilization created for computational purposes only =====
# creates a new_List, will be used to subtract original list (same size)
# removes 1st elemement, shifts 1 left, adds final value to last element
#   input:
#       List : original list
#   output:
#       new_List: will be used to subtract original list (same size)
# ideal
comp_utilization_Add_CC = shift_one_add_final (utilization_Add_CC)
comp_utilization_Add_CT = shift_one_add_final (utilization_Add_CT)
comp_utilization_Add_RIC = shift_one_add_final (utilization_Add_RIC)
comp_utilization_Add_ARO = shift_one_add_final (utilization_Add_ARO)
# choice 1
comp_utilization_Add1_CC = shift_one_add_final (utilization_Add1_CC)
comp_utilization_Add1_CT = shift_one_add_final (utilization_Add1_CT)
comp_utilization_Add1_RIC = shift_one_add_final (utilization_Add1_RIC)
comp_utilization_Add1_ARO = shift_one_add_final (utilization_Add1_ARO)

#print ("\ncomp_utilization_Add_CC: ",comp_utilization_Add_CC[nAddCC1])
#print ("len(comp_utilization_Add_CC): ", len(comp_utilization_Add_CC))
# ==========================================================================
#                           Derivative
# finds the slope derivative of a list
# used for 1st & 2nd derivative
#   input:
#       STEP : Step size
#       List: original list
#       new_List: list used to subtract original list
#   output:
#       derivat: derivative list size 
# ideal
Add_CC_FstDeriv = derivative (STEP, utilization_Add_CC,comp_utilization_Add_CC )
Add_CC_comp_FstDeriv= shift_one_add_final (Add_CC_FstDeriv )
Add_CC_SecDeriv= derivative (STEP, Add_CC_FstDeriv ,Add_CC_comp_FstDeriv )

Add_CT_FstDeriv = derivative (STEP, utilization_Add_CT,comp_utilization_Add_CT )
Add_CT_comp_FstDeriv= shift_one_add_final (Add_CT_FstDeriv )
Add_CT_SecDeriv= derivative (STEP, Add_CT_FstDeriv ,Add_CT_comp_FstDeriv )

Add_RIC_FstDeriv = derivative (STEP, utilization_Add_RIC,comp_utilization_Add_RIC )
Add_RIC_comp_FstDeriv= shift_one_add_final (Add_RIC_FstDeriv )
Add_RIC_SecDeriv= derivative (STEP, Add_RIC_FstDeriv ,Add_RIC_comp_FstDeriv )

Add_ARO_FstDeriv = derivative (STEP, utilization_Add_ARO,comp_utilization_Add_ARO )
Add_ARO_comp_FstDeriv= shift_one_add_final (Add_ARO_FstDeriv )
Add_ARO_SecDeriv= derivative (STEP, Add_ARO_FstDeriv ,Add_ARO_comp_FstDeriv )

# choice 1
Add1_CC_FstDeriv = derivative (STEP, utilization_Add1_CC,comp_utilization_Add1_CC )
Add1_CC_comp_FstDeriv= shift_one_add_final (Add1_CC_FstDeriv )
Add1_CC_SecDeriv= derivative (STEP, Add1_CC_FstDeriv ,Add1_CC_comp_FstDeriv )

Add1_CT_FstDeriv = derivative (STEP, utilization_Add1_CT,comp_utilization_Add1_CT )
Add1_CT_comp_FstDeriv= shift_one_add_final (Add1_CT_FstDeriv )
Add1_CT_SecDeriv= derivative (STEP, Add1_CT_FstDeriv ,Add1_CT_comp_FstDeriv )

Add1_RIC_FstDeriv = derivative (STEP, utilization_Add1_RIC,comp_utilization_Add1_RIC )
Add1_RIC_comp_FstDeriv= shift_one_add_final (Add1_RIC_FstDeriv )
Add1_RIC_SecDeriv= derivative (STEP, Add1_RIC_FstDeriv ,Add1_RIC_comp_FstDeriv )

Add1_ARO_FstDeriv = derivative (STEP, utilization_Add1_ARO,comp_utilization_Add1_ARO )
Add1_ARO_comp_FstDeriv= shift_one_add_final (Add1_ARO_FstDeriv )
Add1_ARO_SecDeriv= derivative (STEP, Add1_ARO_FstDeriv ,Add1_ARO_comp_FstDeriv )


# -----for best choice only examines from 0 to .9 * N (total number of steps)---
# ------------------- when END = 2 * nameplate -------------------------------
# ---------- the rest of the 2nd derivitate is filled with 0 -----------------
# ----------------- prvents calulational errors -----------------------------
#   input:
#       Sec_derivat: the list [...] with 2nd derivatives
#   output:
#       add_zeros: last 2 elements in list [sizes] position with [..0,0]
# ideal
Add_CC_SecDeriv= adjSecDeriv (Add_CC_SecDeriv, Test)
Add_CT_SecDeriv= adjSecDeriv (Add_CT_SecDeriv, Test)
Add_RIC_SecDeriv= adjSecDeriv (Add_RIC_SecDeriv, Test)
Add_ARO_SecDeriv= adjSecDeriv (Add_ARO_SecDeriv, Test)
# choice 1
Add1_CC_SecDeriv= adjSecDeriv (Add1_CC_SecDeriv, Test)
Add1_CT_SecDeriv= adjSecDeriv (Add1_CT_SecDeriv, Test)
Add1_RIC_SecDeriv= adjSecDeriv (Add1_RIC_SecDeriv, Test)
Add1_ARO_SecDeriv= adjSecDeriv (Add1_ARO_SecDeriv, Test)


# ===========================================================================
# ----------------------- conbiniation Additional + gas ---------------------
# === Suggests best choice additional resource based on 2nd derivative ======
#   input:
#       START: starting size of farm used in sizes array
#       STEP : step oe increasing farm sizes
#       derivat: list of 2nd derivative [sizes] 
#   output:
#       choice :  suggested farm size  (var)
#       index:    position of farm size in size list (var)
#       sz:       size of the farm recommended (var)

# ideal 
Add_CC_choice, nAddCC, Add_CC_sz = best_choice \
    (START,STEP, Add_CC_SecDeriv, ((wind_Id or solar_Id) * CC_Id), nAdd)  
Add_CT_choice, nAddCT, Add_CT_sz = best_choice \
    (START,STEP, Add_CT_SecDeriv, ((wind_Id or solar_Id) * CT_Id), nAdd)  
Add_RIC_choice, nAddRIC, Add_RIC_sz = best_choice \
    (START,STEP, Add_RIC_SecDeriv, ((wind_Id or solar_Id) * RIC_Id), nAdd)  
Add_ARO_choice, nAddARO, Add_ARO_sz = best_choice \
    (START,STEP, Add_ARO_SecDeriv, ((wind_Id or solar_Id) * ARO_Id), nAdd)
# choice 1
Add1_CC_choice, nAddCC1, Add1_CC_sz = best_choice \
    (START,STEP, Add1_CC_SecDeriv, (chs_add1 * CC_Id), nAdd1)  
Add1_CT_choice, nAddCT1, Add1_CT_sz = best_choice \
    (START,STEP, Add1_CT_SecDeriv, (chs_add1 * CT_Id), nAdd1)  
Add1_RIC_choice, nAddRIC1, Add1_RIC_sz = best_choice \
    (START,STEP, Add1_RIC_SecDeriv, (chs_add1 * RIC_Id), nAdd1)  
Add1_ARO_choice, nAddARO1, Add1_ARO_sz = best_choice \
    (START,STEP, Add1_ARO_SecDeriv, (chs_add1 * ARO_Id), nAdd1)
    
# ----------------------------------------------------------------------------
# single resources
# ideal
print("ideal, single resources: ")
print ("Additional_choice: ",round (Additional_choice,8), ", nAdd: ",nAdd, ", \
          Add_sz: ", Add_sz)
print ("CC_choice: ", round (CC_choice,8), " , nCC: ",nCC, " , CC_sz: ", CC_sz)
print ("CT_choice: ", round (CT_choice,8), " , nCT: ",nCT, " , CT_sz: ", CT_sz)
print ("RIC_choice: ", round (RIC_choice,8), ", nRIC: ",nRIC, ", RIC_sz: ", RIC_sz)
print ("ARO_choice: ", round (ARO_choice,8), ", nARO: ",nARO, ", ARO_sz: ", ARO_sz)

# Batteries -combination     
# picking batery size (3 user choices) code determines ideal additional size 
print("ideal, combination battery: ") 
print ("AddBat1_choice: ", round (AddBat1_choice,8), " ,  nAddBat1: ", nAddBat1, \
       " , AddBat1_sz: ", AddBat1_sz)
print ("AddBat2_choice: ", round (AddBat2_choice,8), " ,  nAddBat2: ", nAddBat2, \
       " , AddBat2_sz: ", AddBat2_sz)
print ("AddBat3_choice: ", round (AddBat3_choice,8), " ,  nAddBat3: ", nAddBat3, \
       " , AddBat3_sz: ", AddBat3_sz)
    
# combination
# ideal (both)
print("\nideal, conbination: ")
print ("Add_CC_choice: ", round (Add_CC_choice,8), " ,  nAddCC: ", nAddCC,\
       " , Add_CC_sz: ", Add_CC_sz)
print ("Add_CT_choice: ", round (Add_CT_choice,8), " ,  nAddCT: ", nAddCT,\
       " , Add_CT_sz: ", Add_CT_sz)
print ("Add_RIC_choice: ", round (Add_RIC_choice,8), ",  nAddRIC: ", nAddRIC,\
       ", Add_RIC_sz: ", Add_RIC_sz)
print ("Add_ARO_choice: ", round (Add_ARO_choice,8), ",  nAddARO: ", nAddARO,\
       ", Add_ARO_sz: ", Add_ARO_sz)

# choice 1 (choose addditionalsize , code picks ideal size gas)
print("\nchoice 1, conbination: ")
print ("Add1_CC_choice: ", round (Add1_CC_choice,8), " , nAddCC1: ", nAddCC1,\
       " , Add1_CC_sz: ", Add1_CC_sz)
print ("Add1_CT_choice: ", round (Add1_CT_choice,8), " , nAddCT1: ", nAddCT1,\
       " , Add1_CT_sz: ", Add1_CT_sz)
print ("Add1_RIC_choice: ", round (Add1_RIC_choice,8), ", nAddRIC1: ", nAddRIC1,\
       ", Add1_RIC_sz: ", Add1_RIC_sz)
print ("Add1_ARO_choice: ", round (Add1_ARO_choice,8), ", nAddARO1: ", nAddARO1,\
       ", Add1_ARO_sz: ", Add1_ARO_sz)
# the rest of the choices (single or combination) are not here, since they do 
# not got through best_choice fn (*are choicen by the user)

# @@@@@@@@@@@@@@@@@ Comb. (Add_ideal + gas_idea) @@@@@@@@@@@ M3 @@@@@@@@@@@@@
# @@@@@@@@@@@@@@@@@ Comb. (Add_1 + gas_idea)     @@@@@@@@@@@ M3 @@@@@@@@@@@@@ 
# @@@@@@@@@@@@@@@@@ Comb. (Add_ideal + gas_1)    @@@@@@@@@@@ M3 @@@@@@@@@@@@@
# @@@@@@@@@@@@@@@@@ Comb. (Add_1 + gas_2)        @@@@@@@@@@@ M3 @@@@@@@@@@@@@      
#----- Gas Total Energy in [MWh] after best choice for Exist + Additional ---
#            determines the total Energy produced each month  
# input:
#   gas_energy_adj: the energy of gas as adjusted to full forcast and MP
# output :
#  Total energy output per month [MWh], var, based on generator size

# for ideal size both (code recomendation)
JrCc, FbCc, MrCc, ApCc, MyCc, JnCc, JlCc, AgCc, SpCc, OcCc, NvCc,\
    DcCc = monthEnergyGas (AddCC_gas_energy_adj,nAddCC) 
JrCt, FbCt, MrCt, ApCt, MyCt, JnCt, JlCt, AgCt, SpCt, OcCt, NvCt,\
    DcCt = monthEnergyGas (AddCT_gas_energy_adj,nAddCT)
JrRic, FbRic, MrRic, ApRic, MyRic, JnRic, JlRic, AgRic, SpRic, OcRic, NvRic,\
    DcRic = monthEnergyGas (AddRIC_gas_energy_adj,nAddRIC) 
Jr_a, Fb_a, Mr_a, Ap_a, My_a, Jn_a, Jl_a, Ag_a, Sp_a, Oc_a, Nv_a,\
    Dc_a = monthEnergyGas (AddARO_gas_energy_adj,nAddARO)

# for user choice 1 (user chooces additional, gas = ideal)
JrCc1, FbCc1, MrCc1, ApCc1, MyCc1, JnCc1, JlCc1, AgCc1, SpCc1, OcCc1, NvCc1,\
    DcCc1 = monthEnergyGas (AddCC1_gas_energy_adj,nAddCC1) 
JrCt1, FbCt1, MrCt1, ApCt1, MyCt1, JnCt1, JlCt1, AgCt1, SpCt1, OcCt1, NvCt1,\
    DcCt1 = monthEnergyGas (AddCT1_gas_energy_adj,nAddCT1)
JrRic1, FbRic1, MrRic1, ApRic1, MyRic1, JnRic1, JlRic1, AgRic1, SpRic1, OcRic1, NvRic1,\
    DcRic1 = monthEnergyGas (AddRIC1_gas_energy_adj,nAddRIC1) 
Jr_a1, Fb_a1, Mr_a1, Ap_a1, My_a1, Jn_a1, Jl_a1, Ag_a1, Sp_a1, Oc_a1, Nv_a1,\
    Dc_a1 = monthEnergyGas (AddARO1_gas_energy_adj,nAddARO1)

# for user choice 2 ( additional= ideal, user chooces gas )
JrCc2, FbCc2, MrCc2, ApCc2, MyCc2, JnCc2, JlCc2, AgCc2, SpCc2, OcCc2, NvCc2,\
    DcCc2 = monthEnergyGas (AddCC_gas_energy_adj,nAddCC2) 
JrCt2, FbCt2, MrCt2, ApCt2, MyCt2, JnCt2, JlCt2, AgCt2, SpCt2, OcCt2, NvCt2,\
    DcCt2 = monthEnergyGas (AddCT_gas_energy_adj,nAddCT2)
JrRic2, FbRic2, MrRic2, ApRic2, MyRic2, JnRic2, JlRic2, AgRic2, SpRic2, OcRic2, NvRic2,\
    DcRic2 = monthEnergyGas (AddRIC_gas_energy_adj,nAddRIC2) 
Jr_a2, Fb_a2, Mr_a2, Ap_a2, My_a2, Jn_a2, Jl_a2, Ag_a2, Sp_a2, Oc_a2, Nv_a2,\
    Dc_a2 = monthEnergyGas (AddARO_gas_energy_adj,nAddARO2)

# for user choice 3 ( user chooces additional, user chooces gas )
JrCc3, FbCc3, MrCc3, ApCc3, MyCc3, JnCc3, JlCc3, AgCc3, SpCc3, OcCc3, NvCc3,\
    DcCc3 = monthEnergyGas (AddCC1_gas_energy_adj,nAddCC3) 
JrCt3, FbCt3, MrCt3, ApCt3, MyCt3, JnCt3, JlCt3, AgCt3, SpCt3, OcCt3, NvCt3,\
    DcCt3 = monthEnergyGas (AddCT1_gas_energy_adj,nAddCT3)
JrRic3, FbRic3, MrRic3, ApRic3, MyRic3, JnRic3, JlRic3, AgRic3, SpRic3, OcRic3, NvRic3,\
    DcRic3 = monthEnergyGas (AddRIC1_gas_energy_adj,nAddRIC3) 
Jr_a3, Fb_a3, Mr_a3, Ap_a3, My_a3, Jn_a3, Jl_a3, Ag_a3, Sp_a3, Oc_a3, Nv_a3,\
    Dc_a3 = monthEnergyGas (AddARO1_gas_energy_adj,nAddARO3)

#^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^    

# &&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&
# ^^^^^^^^^^^^^^^^^^^^^^end^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^

# @@@@@@@@@@@@@@@@@@@@ plots @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

# === Suggests best choice additional resource based on 2nd derivative ======
#   input:
#       lots for picking up the NAMES for the names & values to be ploted
#       the str(resource) & LABEL for xlabel, ylabel, title 
#   output:
#       name, value, xlabel , ylabel, title
# resources = ['Added', 'CC', 'CT', 'RIC', 'ARO', \
#             'Add_battery','Add_CC', 'Add_CT', 'Add_RIC', 'Add_ARO']
def naming_plot (resource, sizes, Sum_added, Sum_curtail, utilization_act,\
         FstDeriv, SecDeriv, var_x, var_y  ):
    NAMES = [sizes, Sum_added, Sum_curtail, utilization_act,\
            FstDeriv, SecDeriv  ]  
    LABEL = [ str(resource)+'_sizes', str(resource)+'_Sum_added',\
          str(resource)+'_Sum_curtail', str(resource)+'_utilization_act',\
            str(resource) +'_FstDeriv', str(resource) +'_SecDeriv']

    name = NAMES[var_x]
    value = NAMES[var_y]
    xlabel = LABEL[var_x]
    ylabel= LABEL[var_y]
    title = [ylabel, ' vs ',xlabel] 
    return name, value, xlabel , ylabel, title

# function turning COSTANT in tp LIST
def const_to_list (const, listOfSz):
    List = []
    for i in range(len(listOfSz)):
        List.append(const)
    return List
# gas curtailment CONST in to LIST
CC_cutr_List =  const_to_list (0, CC_sizes)
CT_cutr_List =  const_to_list (0, CC_sizes)
RIC_cutr_List = const_to_list (0, CC_sizes)
ARO_cutr_List = const_to_list (0, CC_sizes)
CC_Comb_cutr_List =   const_to_list (Sum_ideal_curtail, CC_sizes)
CT_Comb_cutr_List =   const_to_list (Sum_ideal_curtail, CC_sizes)
RIC_Comb_cutr_List =  const_to_list (Sum_ideal_curtail, CC_sizes)
ARO_Comb_cutr_List =  const_to_list (Sum_ideal_curtail, CC_sizes)

# resources = [0.'Added', 1.'Add_battery', \
#            2.'CC', 3.'CT', 4.'RIC', 5.'ARO', 
#               6.'Add_CC', 7.'Add_CT', 8.'Add_RIC', 9.'Add_ARO']

# piking up the specifices for Solar & gas
x_Additional, y_Additional, Additional_xlabel , Additional_ylabel, Additional_title =\
    naming_plot (resources[0], Additional_sizes, Sum_Add_added, Sum_Addition_curtail, \
         utilization_Add, Additional_FstDeriv, Additional_SecDeriv, \
             Additional_var_x, Additional_var_y )
# -----------------------------------------------------------------------------
x_CC, y_CC, CC_xlabel , CC_ylabel, CC_title =\
    naming_plot (resources[1], CC_sizes, Sum_CC_added, CC_cutr_List, \
         utilization_CC, CC_FstDeriv, CC_SecDeriv, \
             CC_var_x, CC_var_y )

x_CT, y_CT, CT_xlabel , CT_ylabel, CT_title =\
    naming_plot (resources[2], CT_sizes, Sum_CT_added, CT_cutr_List, \
         utilization_CT, CT_FstDeriv, CT_SecDeriv, \
             CT_var_x, CT_var_y )
        
x_RIC, y_RIC, RIC_xlabel , RIC_ylabel, RIC_title =\
    naming_plot (resources[3], RIC_sizes, Sum_RIC_added, RIC_cutr_List, \
         utilization_RIC, RIC_FstDeriv, RIC_SecDeriv, \
             RIC_var_x, RIC_var_y )
        
x_ARO, y_ARO, ARO_xlabel , ARO_ylabel, ARO_title =\
    naming_plot (resources[4], ARO_sizes, Sum_ARO_added, ARO_cutr_List, \
         utilization_ARO, ARO_FstDeriv, ARO_SecDeriv, \
             ARO_var_x, ARO_var_y )
# -----------------------------------------------------------------------------
                                                        
x_AddBat2, y_AddBat2, AddBat2_xlabel , AddBat2_ylabel, AddBat2_title =\
    naming_plot (resources[5], Additional_sizes, AddBat3_tot, AddBat3_curtail, \
         utilization_AddBat3, AddBat3_FstDeriv, AddBat3_SecDeriv, \
             AddBat_var_x, AddBat_var_y )    
# ---------------------------------------------------------------------------        
x_Add_CC, y_Add_CC, Add_CC_xlabel , Add_CC_ylabel, Add_CC_title =\
    naming_plot (resources[6], Add_CC_sizes, Sum_Add_CC_added, CC_Comb_cutr_List, \
         utilization_Add_CC, Add_CC_FstDeriv, Add_CC_SecDeriv, \
             Add_CC_var_x, Add_CC_var_y )

x_Add_CT, y_Add_CT, Add_CT_xlabel , Add_CT_ylabel, Add_CT_title =\
    naming_plot (resources[7], Add_CT_sizes, Sum_Add_CT_added, CT_Comb_cutr_List, \
         utilization_Add_CT, Add_CT_FstDeriv, Add_CT_SecDeriv, \
             Add_CT_var_x, Add_CT_var_y )
        
x_Add_RIC, y_Add_RIC, Add_RIC_xlabel , Add_RIC_ylabel, Add_RIC_title =\
    naming_plot (resources[8], Add_RIC_sizes, Sum_Add_RIC_added, RIC_Comb_cutr_List, \
         utilization_Add_RIC, Add_RIC_FstDeriv, Add_RIC_SecDeriv, \
             Add_RIC_var_x, Add_RIC_var_y )
        
x_Add_ARO, y_Add_ARO, Add_ARO_xlabel , Add_ARO_ylabel, Add_ARO_title =\
    naming_plot (resources[9], Add_ARO_sizes, Sum_Add_ARO_added, ARO_Comb_cutr_List, \
         utilization_Add_ARO, Add_ARO_FstDeriv, Add_ARO_SecDeriv, \
             Add_ARO_var_x, Add_ARO_var_y )
        

# =====================  line graph ==========================================
# === Creating the line graph  =================================
#   input:
#       name, value, xlabel , ylabel, title
#   output:
#       graph

def line_graph (name, value, xlabel , ylabel, title ):
    #for i in range(len(name)):
    plt.plot(name, value)
     
    plt.title(title)
    plt.xlabel(xlabel )
    plt.ylabel(ylabel)
    graph =plt.show()
    return graph

Additional_graph =  line_graph (x_Additional,\
    y_Additional, Additional_xlabel , Additional_ylabel, Additional_title )
CC_graph =  line_graph (x_CC, y_CC, CC_xlabel , CC_ylabel, CC_title )
CT_graph =  line_graph (x_CT, y_CT, CT_xlabel , CT_ylabel, CT_title )
RIC_graph =  line_graph (x_RIC, y_RIC, RIC_xlabel , RIC_ylabel, RIC_title )
ARO_graph =  line_graph (x_ARO, y_ARO, ARO_xlabel , ARO_ylabel, ARO_title )

AddBat2_graph =  line_graph \
    (x_AddBat2, y_AddBat2, AddBat2_xlabel , AddBat2_ylabel, AddBat2_title )
Add_CC_graph =  line_graph \
    (x_Add_CC, y_Add_CC, Add_CC_xlabel , Add_CC_ylabel, Add_CC_title )
Add_CT_graph =  line_graph \
    (x_Add_CT, y_Add_CT, Add_CT_xlabel , Add_CT_ylabel, Add_CT_title )
Add_RIC_graph =  line_graph \
    (x_Add_RIC, y_Add_RIC, Add_RIC_xlabel , Add_RIC_ylabel, Add_RIC_title )
Add_ARO_graph =  line_graph \
    (x_Add_ARO, y_Add_ARO, Add_ARO_xlabel , Add_ARO_ylabel, Add_ARO_title )

# @@@@@@@@@@@@@@@@@@@@@@@@ end ploting @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@


# $$$$$$$$$$$$$$$$$$$$$$$$ MODULE 3 $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$

#-----------------------------Input from M2------------------------------------
# This is a list [sizes], var = list [n]

#                 Energy [MWh] & cutrilment ideal
# for 'ideal resources' use if...then to implement user choices 
def user_decisn_idl ( user_choice, position):
    if user_choice == True:
        MWhAddtn= Sum_Add_added[position]              # Energy ideal code chosen 
        MWhAddtn_Tcrt = Sum_Addition_curtail[position] # curtailment ideal code chosen 
    else:
        MWhAddtn= 0        # Energy ideal code chosen 
        MWhAddtn_Tcrt = 0  # curtailment ideal code chosen 
    return MWhAddtn, MWhAddtn_Tcrt
MWhAddtn, MWhAddtn_Tcrt = user_decisn_idl ( wind_Id | solar_Id, nAdd)

#                           Added Energy [MWh]
# [MWh] in Year  'Additional resource' added (alone or Combination) // wind or solar
# var = list[sizes]
# Added additional [MWh]
MWhAddtn1= Sum_Add_added[nAdd1]     # used choice 1 (additional resource alone)
MWhAddtn2= Sum_Add_added[nAdd2]     # used choice 2 (additional resource alone)

#                   Energy Curtailment - non ideal
# [MWh] Total (used & stored) - Solar (use when NO battery)
# var,  # curtailment sum for a size of Additional farm/(sum of 8760)  
# gas produces no curtailment
MWhAddtn1_Tcrt = Sum_Addition_curtail[nAdd1] # choice 1, curtailment sum 
MWhAddtn2_Tcrt = Sum_Addition_curtail[nAdd2] # choice 2, curtailment sum

# ------------------------------- battery --------------------------------------
# battery , Addtional sz ideal

#                 Energy [MWh] & cutrilment ideal
# for 'ideal resources' use if...then to implement user choices 
def user_decisn_bat ( user_choice, Energy, Curtailment ):
    if user_choice == True:
        MWhAddtn= Energy            # Energy ideal code chosen 
        MWhAddtn_Tcrt = Curtailment # curtailment ideal code chosen 
    else:
        MWhAddtn= 0        # Energy ideal code chosen 
        MWhAddtn_Tcrt = 0  # curtailment ideal code chosen 
    return MWhAddtn, MWhAddtn_Tcrt

# battery Energy & Cutr - the nAddBat1 = position in additional farm list
#                                        based on battery size
MWhAddtn1_bat, MWhAddtnBat1_Tcrt = \
    user_decisn_bat (Bat1_Id and chs_Bat1, AddBat1_tot[nAddBat1], \
                     AddBat1_curtail[nAddBat1])
MWhAddtn2_bat, MWhAddtnBat2_Tcrt = \
    user_decisn_bat (Bat1_Id and chs_Bat2, AddBat2_tot[nAddBat2], \
                     AddBat2_curtail[nAddBat2])
MWhAddtn3_bat, MWhAddtnBat3_Tcrt = \
    user_decisn_bat (Bat1_Id and chs_Bat3, AddBat3_tot[nAddBat3], \
                     AddBat3_curtail[nAddBat3])

# @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

# ************************** openpyxl ***************************************  
#--FORMATING using 'openpyxl' copy from 'M3OutputExample.xlsx' to 'ProjectResults.xlsx'---
# Excel-Python using 'openpyxl'

# ----------------'WRITING' on Excel 'EcoAnResuls.xlsx'----------------------
# create blank Workbook object & active sheet 
wbd = openpyxl.Workbook()     # wb destination
page=wbd.active               # page @ destination
# page titles:
page.title = 'EcoAnResuls'

#-------load 'M3OutputExample.xlsx'---------------------------------------------
wbs2= openpyxl.load_workbook('M3OutputExample.xlsx')
sheet2 = wbs2.active 

#---from 'M3OutputExample.xlsx' to 'EcoAnResuls.xlsx'--------------------------
# total number of rows & columns in 'M3OutputExample.xlsx'----------------
Rm = sheet2.max_row 
Cm = sheet2.max_column 
  
# copying the cell values from source 'M3OutputExample.xlsx' 
for i in range (1, Rm + 1): 
    for j in range (1, Cm + 1):  
        Cell =sheet2.cell(row = i, column = j).value 
  
        # Transfer value to destination 'EcoAnResuls.xlsx' 
        page.cell(row = i, column = j).value = Cell 
#-------------------------------FORMATTING------------------------------------
# * wider cells


page.column_dimensions['C'].width = 30
page.column_dimensions['D'].width = 20
page.column_dimensions['E'].width = 20

# Bold --------------------------------------------------
Bold = Font(bold=True)
page['A1'].font = Bold
page['C1'].font = Bold
page['A2'].font = Bold
page['A44'].font = Bold
page['C44'].font = Bold
page['A45'].font = Bold
# Merge--------------------------------------------------
page.merge_cells('D2:G2')
page.merge_cells('D45:G45')
# 2 decimal places----------------------------------------
for row in range(1,Rm+1):
    page["D{}".format(row)].number_format = '#,##0.00'
    page["E{}".format(row)].number_format = '#,##0.00'
    page["F{}".format(row)].number_format = '#,##0.00'
    page["G{}".format(row)].number_format = '#,##0.00'
    page["H{}".format(row)].number_format = '#,##0.00'
    page["I{}".format(row)].number_format = '#,##0.00'
# Wrap_text & Center-----------------------------------------
for row in page.iter_rows():
    for cell in row:      
        alignment = copy.copy(cell.alignment)
        alignment.wrapText=True
        alignment.horizontal = 'center'
        alignment.vertical = 'center'
        cell.alignment = alignment
# Border------------------------------------------------------
LeftThck = Border(left=Side(style='thick'))    # col D
RightThck = Border(right=Side(style='thick'))  # col I
TopThck = Border(top=Side(style='thick'))      # row 38,81
BottThck = Border(bottom=Side(style='thick'))  # row 3,46
# row for existing Wind                     
for row in page.iter_rows(min_row=4, max_row=37, min_col=4, max_col=4):
    for cell in row:
        cell.border=LeftThck
for row in page.iter_rows(min_row=4, max_row=37, min_col=9, max_col=9):
    for cell in row:
        cell.border=RightThck
for row in page.iter_rows(min_row=3, max_row=3, min_col=4, max_col=9):
    for cell in row:
        cell.border=BottThck
for row in page.iter_rows(min_row=38, max_row=38, min_col=4, max_col=9):
    for cell in row:
        cell.border=TopThck
# row for existing Solar                     
for row in page.iter_rows(min_row=47, max_row=80, min_col=4, max_col=4):
    for cell in row:
        cell.border=LeftThck
for row in page.iter_rows(min_row=47, max_row=80, min_col=9, max_col=9):
    for cell in row:
        cell.border=RightThck
for row in page.iter_rows(min_row=46, max_row=46, min_col=4, max_col=9):
    for cell in row:
        cell.border=BottThck
for row in page.iter_rows(min_row=81, max_row=81, min_col=4, max_col=9):
    for cell in row:
        cell.border=TopThck
        
# ********************* End formating openpyxl *******************************  

#************************* Actual Code Starts here ***************************
#------------------ List yearMWh =[monthMWh,monthMWh... ] --------------------
# --------------------- Single gas ONLY additional resource --------------------
# -------------------------------- ideal --------------------------------------
# CC MWh in Year
MWhCC =[JrCC,FbCC,MrCC,ApCC,MyCC,JnCC,         # this is the val. of the list
        JlCC,AgCC,SpCC,OcCC,NvCC,DcCC]
# CT MWh in Year
MWhCT =[JrCT,FbCT,MrCT,ApCT,MyCT,JnCT,         # this is the val. of the list
        JlCT,AgCT,SpCT,OcCT,NvCT,DcCT]
# RIC MWh in Year
MWhRIC =[JrRIC,FbRIC,MrRIC,ApRIC,MyRIC,JnRIC,  # this is the val. of the list
        JlRIC,AgRIC,SpRIC,OcRIC,NvRIC,DcRIC]
# A MWh in Year
MWhA =[JrA,FbA,MrA,ApA,MyA,JnA,                # this is the val. of the list
        JlA,AgA,SpA,OcA,NvA,DcA]

# -------------------------------- choice 1 ------------------------------------
MWhCC1 =[JrCC1, FbCC1, MrCC1, ApCC1, MyCC1, JnCC1,       # val. of the list
        JlCC1, AgCC1, SpCC1, OcCC1, NvCC1, DcCC1]
# CT MWh in Year
MWhCT1 =[JrCT1, FbCT1, MrCT1, ApCT1, MyCT1, JnCT1,       # val. of the list
        AgCT1, SpCT1, OcCT1, NvCT1, DcCT1]
# RIC MWh in Year
MWhRIC1 =[JrRIC1, FbRIC1, MrRIC1, ApRIC1, MyRIC1, JnRIC1, # val. of the list
        JlRIC1, AgRIC1, SpRIC1, OcRIC1, NvRIC1, DcRIC1]
# A MWh in Year
MWhA1 =[JrA1, FbA1, MrA1, ApA1, MyA1, JnA1,               # val. of the list
        JlA1, AgA1, SpA1, OcA1, NvA1, DcA1]

# -------------------------------- choice 2 -----------------------------------
MWhCC2 =[JrCC2, FbCC2, MrCC2, ApCC2, MyCC2, JnCC2,        # val. of the list
        JlCC2, AgCC2, SpCC2, OcCC2, NvCC2, DcCC2]
# CT MWh in Year
MWhCT2 =[JrCT2, FbCT2, MrCT2, ApCT2, MyCT2, JnCT2,        # val. of the list
        JlCT2, AgCT2, SpCT2, OcCT2, NvCT2, DcCT2]
# RIC MWh in Year
MWhRIC2 =[JrRIC2, FbRIC2, MrRIC2, ApRIC2, MyRIC2, JnRIC2, # val. of the list
        JlRIC2, AgRIC2, SpRIC2, OcRIC2, NvRIC2, DcRIC2]
# A MWh in Year
MWhA2 =[JrA2, FbA2, MrA2, ApA2, MyA2, JnA2,               # val. of the list
        JlA2, AgA2, SpA2, OcA2, NvA2, DcA2 ]

# =============== When gas COMBINED additional resource ======================
# ------------------ Comb. (Add_ideal + gas_idea) ----------------------------
# Cc MWh in Year
MWhCc =[JrCc,FbCc,MrCc,ApCc,MyCc,JnCc,         # this is the val. of the list
        JlCc,AgCc,SpCc,OcCc,NvCc,DcCc]
# Ct MWh in Year
MWhCt =[JrCt,FbCt,MrCt,ApCt,MyCt,JnCt,         # this is the val. of the list
        JlCt,AgCt,SpCt,OcCt,NvCt,DcCt]
# Ric MWh in Year
MWhRic =[JrRic,FbRic,MrRic,ApRic,MyRic,JnRic,  # this is the val. of the list
        JlRic,AgRic,SpRic,OcRic,NvRic,DcRic]
# _a MWh in Year
MWh_a =[Jr_a,Fb_a,Mr_a,Ap_a,My_a,Jn_a,         # this is the val. of the list
        Jl_a,Ag_a,Sp_a,Oc_a,Nv_a,Dc_a]

#print('MWh_a \n', MWh_a)
# ------------------ choice 1, Comb. (Add_1 + gas_idea) -----------------------
# Cc MWh in Year
MWhCc1 =[JrCc1, FbCc1, MrCc1, ApCc1, MyCc1, JnCc1,        # val. of the list
        JlCc1, AgCc1, SpCc1, OcCc1, NvCc1 ,DcCc1]
# Ct MWh in Year
MWhCt1 =[JrCt1, FbCt1, MrCt1, ApCt1, MyCt1, JnCt1,        # val. of the list
        JlCt1, AgCt1, SpCt1, OcCt1, NvCt1, DcCt1 ]
# Ric MWh in Year
MWhRic1 =[JrRic1, FbRic1, MrRic1, ApRic1, MyRic1, JnRic1, # val. of the list
        JlRic1, AgRic1, SpRic1, OcRic1, NvRic1, DcRic1 ]
# _a MWh in Year
MWh_a1 =[Jr_a1, Fb_a1, Mr_a1, Ap_a1, My_a1, Jn_a1,       # val. of the list
        Jl_a1, Ag_a1, Sp_a1, Oc_a1, Nv_a1, Dc_a1 ]

# ------------------ choice 2, Comb. (Add_ideal + gas_1)  ---------------------
# Cc MWh in Year
MWhCc2 =[JrCc2, FbCc2, MrCc2, ApCc2, MyCc2, JnCc2,        # val. of the list
        JlCc2, AgCc2, SpCc2, OcCc2, NvCc2, DcCc2]
# Ct MWh in Year
MWhCt2 =[JrCt2, FbCt2, MrCt2, ApCt2, MyCt2, JnCt2,        # val. of the list
        JlCt2, AgCt2, SpCt2, OcCt2, NvCt2, DcCt2]
# Ric MWh in Year
MWhRic2 =[JrRic2, FbRic2, MrRic2, ApRic2, MyRic2, JnRic2, # val. of the list
        JlRic2, AgRic2, SpRic2, OcRic2, NvRic2, DcRic2]
# _a MWh in Year
MWh_a2 =[Jr_a2, Fb_a2, Mr_a2, Ap_a2, My_a2, Jn_a2,        # val. of the list
        Jl_a2, Ag_a2, Sp_a2, Oc_a2, Nv_a2, Dc_a2]

# ------------------ choice 3, Comb. (Add_1 + gas_2)   ---------------------
# Cc MWh in Year
MWhCc3 =[JrCc3, FbCc3, MrCc3, ApCc3, MyCc3, JnCc3,       # val. of the list
        JlCc3, AgCc3, SpCc3, OcCc3, NvCc3, DcCc3]
# Ct MWh in Year
MWhCt3 =[JrCt3, FbCt3, MrCt3, ApCt3, MyCt3, JnCt3,        # val. of the list
        JlCt3, AgCt3, SpCt3, OcCt3, NvCt3, DcCt3 ]
#print ('MWhCt3: ', MWhCt3)
# Ric MWh in Year
MWhRic3 =[JrRic3, FbRic3, MrRic3, ApRic3, MyRic3, JnRic3, # val. of the list
        JlRic3, AgRic3, SpRic3, OcRic3, NvRic3, DcRic3]
# _a MWh in Year
MWh_a3 =[Jr_a3, Fb_a3, Mr_a3, Ap_a3, My_a3, Jn_a3,        # val. of the list
        Jl_a3, Ag_a3, Sp_a3, Oc_a3, Nv_a3, Dc_a3 ]
    
#--------------------------Formulas------------------------------------------
# Setting the constants from xlsx
# input:
#    xl1: row
#    xl2: column
# output
#   xlsx : value form xlsx cell

# assign values to var in [$/MWh] from 'EconomicAnalysisM3.xlsx'
C26 = pickFROMxlsx(24, 2)    # $/MWh wind
D27 = pickFROMxlsx(25, 3)    # $/MWh solar

#RESOURCE---RESOURCE---RESOURCE---RESOURCE---RESOURCE---RESOURCE------------- 
#---------------------- Share Senarios (All Existing Resources) ---------------
#----------- Calculating gas values, when gas = Single resource --------------

# pulls the 'forward curve' from 'EconomicAnalysisM3.xlsx' & makes list = []
frw_crv = []
for i in range (27,39):
    x = pickFROMxlsx (i, 1)
    frw_crv.append(x)
#print ("frw_crv: \n", frw_crv)

# ======== COMPLETE COMB. (ADD + COMB) fn ======================================
# can be used by (single, or comb.) gas resources
# inputs
#   CpE = Cost per Energy for Additional resource (wind C26, solar D27)
#   HR = heat rate, diff. for CC, CT, RIC, ARO - var.
#   VOM, diff. for CC, CT, RIC, ARO - var.
#   frw_crv = same for all,- list []
#   MWhAddtn = energy from Additonal resource ( | 1) in comb., single =0 -var
#   MWh_gas = Energy from gas res. speciffic to type & (single vs comb.) - list
#   MWh_Tcrt = Total curtailment, ( | 1) in comb., single =0 -var
# for Single gas use CpE= 0, MWhAddtn = 0, MWh_Tcrt = 0
# @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ 
def gasCombResource (CpE, HR,VOM,frw_crv, MWhAddt, MWh_gas, Tcrt):
    
    # initializing var
    SumMWh_g =0; D_g =0     # SumMWh = 0; D=0; 
    
    D_Add = CpE * MWhAddt   # $  Addtn resource, in a yr
    
    for month in range(len(MWh_gas)):  
        # monthly values of Energy & Cost
        DpMWh= ((frw_crv[month]* HR)/1000) + VOM  # [$/MWH]/month of gas resrc.
        monthD = DpMWh * MWh_gas[month]           # [$]/month of gas resrc.           
        
        # Sum  of monthly values
        SumMWh_g += MWh_gas[month]    # Total [MWH]/yr produced by gas resource
        D_g += monthD                 # Total [$]/yr
    
    # values to be returned
    Curt_D = CpE * Tcrt         # $  Curtailment for Addtn resource, in yr    
    
    if (MWhAddt and SumMWh_g) != 0: # Energy for both then calc.
        SumMWh = MWhAddt + SumMWh_g # Tot. [MWh]/yr produced by (Addtn+gas)
        Tcrt = Tcrt                 # if gas is evaluated Curt = Curt Additional
        D = D_Add + D_g             # Tot. [$]/yr produced by (Addtn+gas)
        DpE = (D+Curt_D)/SumMWh # [$(resources(both+cutr)/[MWh/yr] resource]
    else:            # if there ISN'T energy for gas then Energy,Curt,Cost = 0
        SumMWh = 0                  # Tot. [MWh]/yr produced by (Addtn+gas)
        Tcrt = 0                    # if no gas not Curt from Additional
        D =0                        # Tot. [$]/yr produced by (Addtn+gas)
        DpE=0                   # total [$/MWh/yr]
      
    return SumMWh, Tcrt, D, Curt_D, DpE

# ------ retrive data from Single (gas resources) -------------------
# CC
SumMWhCC, CC_Tcrt, CC_D, CC_Curt_D, CC_DpE = gasCombResource \
    (0, CC_HR,CC_VOM,frw_crv, 1, MWhCC, 0)   # ideal CC 
SumMWhCC1, CC1_Tcrt, CC1_D, CC1_Curt_D, CC1_DpE = gasCombResource \
    (0, CC_HR,CC_VOM,frw_crv, 1, MWhCC1, 0)  # 1. cho1 CC 
SumMWhCC2, CC2_Tcrt, CC2_D, CC2_Curt_D, CC2_DpE = gasCombResource \
    (0, CC_HR,CC_VOM,frw_crv, 1, MWhCC2, 0)  # 2. cho2 CC 
#print ("\nSumMWhCC: ", SumMWhCC, "CC_Tcrt: ", CC_Tcrt, "CC_D: ", CC_D, \
#       "CC_Curt_D: ",CC_Curt_D, "CC_DpE: ",CC_DpE ) 

# CT
SumMWhCT, CT_Tcrt, CT_D, CT_Curt_D, CT_DpE = gasCombResource \
    (0, CT_HR,CT_VOM,frw_crv, 1, MWhCT, 0)   # ideal CC 
SumMWhCT1, CT1_Tcrt, CT1_D, CT1_Curt_D, CT1_DpE = gasCombResource \
    (0, CT_HR,CT_VOM,frw_crv, 1, MWhCT1, 0)  # 1. cho1 CT 
SumMWhCT2, CT2_Tcrt, CT2_D, CT2_Curt_D, CT2_DpE = gasCombResource \
    (0, CT_HR,CT_VOM,frw_crv, 1, MWhCT2, 0)  # 2. cho2 CT 

# RIC
SumMWhRIC, RIC_Tcrt, RIC_D, RIC_Curt_D, RIC_DpE = gasCombResource \
    (0, RIC_HR,RIC_VOM,frw_crv, 1, MWhRIC, 0)   # ideal RIC 
SumMWhRIC1, RIC1_Tcrt, RIC1_D, RIC1_Curt_D, RIC1_DpE = gasCombResource \
    (0, RIC_HR,RIC_VOM,frw_crv, 1, MWhRIC1, 0)  # 1. cho1 RIC 
SumMWhRIC2, RIC2_Tcrt, RIC2_D, RIC2_Curt_D, RIC2_DpE = gasCombResource \
    (0, RIC_HR,RIC_VOM,frw_crv, 1, MWhRIC2, 0)  # 2. cho2 RIC 

# ARO or A
SumMWhA, A_Tcrt, A_D, A_Curt_D, A_DpE = gasCombResource \
    (0, ARO_HR,ARO_VOM,frw_crv, 1, MWhA, 0)   # ideal ARO 
SumMWhA1, A1_Tcrt, A1_D, A1_Curt_D, A1_DpE = gasCombResource \
    (0, ARO_HR,ARO_VOM,frw_crv, 1, MWhA1, 0)  # 1. cho1 ARO 
SumMWhA2, A2_Tcrt, A2_D, A2_Curt_D, A2_DpE = gasCombResource \
    (0, ARO_HR,ARO_VOM,frw_crv, 1, MWhA2, 0)  # 2. cho2 ARO  

Existing_Wind  = 0 & 2 & 3 & 4 & 5 & 6 
Existing_Solar = 1                                                             
                                                                               
# !!!!!!!!!!!!!! Function for Additional | additional +Battery !!!!!!!!!!!!!!!!

# ======== COMPLETE COMB. (ADD + COMB) fn ======================================
# can be used by (single, or comb.) Additional resources
# inputs
#   CpE = Cost per Energy for Additional resource (wind C26, solar D27)
#   MWhAddtn = energy from Additonal resource -var
#   MWh_Tcrt = Total curtailment, ( | 1) in comb., single =0 -var

# 1a) Additional (single) |  Additional +  Battery
def AdditionalBatTF (CpE, MWhAddt, Tcrt):
    
    D = (CpE * MWhAddt)   # $  Addtn + Batt, in a yr
    Curt_D = CpE * Tcrt   # $  Curtailment for Addtn resource, in yr
    
    # Total $/MWh/yr - cannnot divide by 0, so if mechanism
    if MWhAddt > 0:
        DpE = (D+Curt_D)/MWhAddt  # [$(Energy(Add+cutr)/[MWh/yr] Tot_resources]
    else: 
        DpE=0                     # total [$/MWh/yr]
    
    # MWh_add, MWh_Curt, $_add, $_curt, $/MWh/yr_add
    return MWhAddt,  Tcrt, D, Curt_D, DpE

# ///////////// test for percentage of line utilization \\\\\\\\\\\\\\\\\\\\\
# finds line utilization as a %
#           Single gas resources (same for all existing resources)
#   input:
#       Sum_existing : the yearly sum energy of the existing resource [MWh], var
#       Sum_Energy : the yearly sum energy of the existing resource [MWh], var
#       Sum_nameplate : the yearly sum energy of the nameplate [MWh], var
#   output:
#       utilization: % of utilization (var)
def utilizationM3 (Sum_existing, Sum_Energy, Sum_nameplate):
    
    utilization = ((Sum_existing + Sum_Energy)/Sum_nameplate)*100
    return  utilization 

util_CC = utilizationM3 (Sum_Existing, SumMWhCC, Sum_nameplate)
util_CC1 = utilizationM3 (Sum_Existing, SumMWhCC1, Sum_nameplate)
util_CC2 = utilizationM3 (Sum_Existing, SumMWhCC2, Sum_nameplate) 
util_CT = utilizationM3 (Sum_Existing, SumMWhCT, Sum_nameplate)
util_CT1 = utilizationM3 (Sum_Existing, SumMWhCT1, Sum_nameplate)
util_CT2 = utilizationM3 (Sum_Existing, SumMWhCT2, Sum_nameplate)
util_RIC = utilizationM3 (Sum_Existing, SumMWhRIC, Sum_nameplate)
util_RIC1 = utilizationM3 (Sum_Existing, SumMWhRIC1, Sum_nameplate)
util_RIC2 = utilizationM3 (Sum_Existing, SumMWhRIC2, Sum_nameplate)
util_ARO = utilizationM3 (Sum_Existing, SumMWhA, Sum_nameplate)
util_ARO1 = utilizationM3 (Sum_Existing, SumMWhA1, Sum_nameplate)
util_ARO2 = utilizationM3 (Sum_Existing, SumMWhA2, Sum_nameplate)

#////////////////////////////////////\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\

# ############## Existing wind Starts #######################################
if loc == Existing_Wind: # the Additional resource is solar
    
    # the 3 Additional resource senarios 
    MWhS, MWhS_Tcrt, Solar_D, SolarCurt_D, Sol_DpE = AdditionalBatTF \
        (D27, MWhAddtn, MWhAddtn_Tcrt)    # ideal Additional  
    MWhS1, MWhS1_Tcrt, Solar1_D, SolarCurt1_D, Sol1_DpE = AdditionalBatTF \
        (D27, MWhAddtn1, MWhAddtn1_Tcrt)  # 1. cho1 Additional 
    MWhS2, MWhS2_Tcrt, Solar2_D, SolarCurt2_D, Sol2_DpE = AdditionalBatTF \
        (D27, MWhAddtn2, MWhAddtn2_Tcrt)  # 2. cho2 Additional
    #print ("\nMWhS: ", MWhS, "MWhS_Tcrt: ", MWhS_Tcrt, "Solar_D: ", Solar_D, \
    #       "SolarCurt_D: ",SolarCurt_D, "Sol_DpE: ",Sol_DpE ) 
    
    # Comb. Battery 
    # 3 choice user picks battery, code picks ideal Additional size
    MWhBat1, Bat1_Tcrt, Bat1_D, Bat1Curt_D, Bat1_DpE = AdditionalBatTF \
        (D27, MWhAddtn1_bat, MWhAddtnBat1_Tcrt )    # ideal Additional  
    MWhBat2, Bat2_Tcrt, Bat2_D, Bat2Curt_D, Bat2_DpE = AdditionalBatTF \
        (D27, MWhAddtn2_bat, MWhAddtnBat2_Tcrt )  # 1. cho1 Additional 
    MWhBat3, Bat3_Tcrt, Bat3_D, Bat3Curt_D, Bat3_DpE = AdditionalBatTF \
        (D27, MWhAddtn3_bat, MWhAddtnBat3_Tcrt )  # 2. cho2 Additional
   
    # the 4 comb. (Additional resource + gas) senarios
    # gasCombResource (CpE, HR,VOM,frw_crv, MWhAddtn, MWh_gas, MWh_Tcrt)
        
    # CC
    SCcE, SCcCrE, SCcD, SCcCrD, SCcDpE = gasCombResource \
        (D27, CC_HR,CC_VOM,frw_crv, MWhAddtn, MWhCc, MWhAddtn_Tcrt)    # both ideal CC 
    SCcE1, SCcCrE1, SCcD1, SCcCrD1, SCcDpE1 = gasCombResource \
        (D27, CC_HR,CC_VOM,frw_crv, MWhAddtn1, MWhCc1, MWhAddtn1_Tcrt) # 1. cho1, ideal CC 
    SCcE2, SCcCrE2, SCcD2, SCcCrD2, SCcDpE2 = gasCombResource \
        (D27, CC_HR,CC_VOM,frw_crv, MWhAddtn, MWhCc2, MWhAddtn_Tcrt)   # 2. ideal, cho1 CC 
    SCcE3, SCcCrE3, SCcD3, SCcCrD3, SCcDpE3 = gasCombResource \
        (D27, CC_HR,CC_VOM,frw_crv, MWhAddtn1, MWhCc3, MWhAddtn1_Tcrt) # 3. cho1, cho2 CC 
    
    # CT
    SCtE, SCtCrE, SCtD, SCtCrD, SCtDpE = gasCombResource \
        (D27, CT_HR,CT_VOM,frw_crv, MWhAddtn, MWhCt, MWhAddtn_Tcrt)    # both ideal CT 
    SCtE1, SCtCrE1, SCtD1, SCtCrD1, SCtDpE1 = gasCombResource \
        (D27, CT_HR,CT_VOM,frw_crv, MWhAddtn1, MWhCt1, MWhAddtn1_Tcrt) # 1. cho1, ideal CT 
    SCtE2, SCtCrE2, SCtD2, SCtCrD2, SCtDpE2 = gasCombResource \
        (D27, CT_HR,CT_VOM,frw_crv, MWhAddtn, MWhCt2, MWhAddtn_Tcrt)   # 2. ideal, cho1 CT
    SCtE3, SCtCrE3, SCtD3, SCtCrD3, SCtDpE3 = gasCombResource \
        (D27, CT_HR,CT_VOM,frw_crv, MWhAddtn1, MWhCt3, MWhAddtn1_Tcrt) # 3. cho1, cho2 CT 
    
    # RIC 
    SRicE, SRicCrE, SRicD, SRicCrD, SRicDpE = gasCombResource \
        (D27, RIC_HR,RIC_VOM,frw_crv, MWhAddtn, MWhRic, MWhAddtn_Tcrt)   # both ideal RIC 
    SRicE1, SRicCrE1, SRicD1, SRicCrD1, SRicDpE1 = gasCombResource \
        (D27, RIC_HR,RIC_VOM,frw_crv, MWhAddtn1, MWhRic1, MWhAddtn1_Tcrt)# 1. cho1, ideal RIC 
    SRicE2, SRicCrE2, SRicD2, SRicCrD2, SRicDpE2 = gasCombResource \
        (D27, RIC_HR,RIC_VOM,frw_crv, MWhAddtn, MWhRic2, MWhAddtn_Tcrt)  # 2. ideal, cho1 RIC 
    SRicE3, SRicCrE3, SRicD3, SRicCrD3, SRicDpE3 = gasCombResource \
        (D27, RIC_HR,RIC_VOM,frw_crv, MWhAddtn1, MWhRic3, MWhAddtn1_Tcrt)  # 3. cho1, cho2 RIC 
    
    # ARO or A _a
    S_aE, S_aCrE, S_aD, S_aCrD, S_aDpE = gasCombResource \
        (D27, ARO_HR,ARO_VOM,frw_crv, MWhAddtn, MWh_a, MWhAddtn_Tcrt)   # both ideal ARO 
    S_aE1, S_aCrE1, S_aD1, S_aCrD1, S_aDpE1 = gasCombResource \
        (D27, ARO_HR,ARO_VOM,frw_crv, MWhAddtn1, MWh_a1, MWhAddtn1_Tcrt)# 1. cho1, ideal ARO 
    S_aE2, S_aCrE2, S_aD2, S_aCrD2, S_aDpE2 = gasCombResource \
        (D27, ARO_HR,ARO_VOM,frw_crv, MWhAddtn, MWh_a2, MWhAddtn_Tcrt)  # 2. ideal, cho1 ARO 
    S_aE3, S_aCrE3, S_aD3, S_aCrD3, S_aDpE3 = gasCombResource \
        (D27, ARO_HR,ARO_VOM,frw_crv, MWhAddtn1, MWh_a3, MWhAddtn1_Tcrt)  # 3. cho1, cho2 ARO  
    
    # ===========================================================================
    # ///////////// test for percentage of line utilization \\\\\\\\\\\\\\\\\\\\\
    # finds line utilization as a %
    # Additional (alone or in combination) resources (depends on existing resource)
    util_Add = utilizationM3 (Sum_Existing, MWhS, Sum_nameplate)
    util_Add1 = utilizationM3 (Sum_Existing, MWhS1, Sum_nameplate)
    util_Add2 = utilizationM3 (Sum_Existing, MWhS2, Sum_nameplate)
    util_Bat1 = utilizationM3 (Sum_Existing, MWhBat1, Sum_nameplate)
    util_Bat2 = utilizationM3 (Sum_Existing, MWhBat2, Sum_nameplate)
    util_Bat3 = utilizationM3 (Sum_Existing, MWhBat3, Sum_nameplate)
    util_AddCC = utilizationM3 (Sum_Existing, SCcE, Sum_nameplate)
    util_AddCC1 = utilizationM3 (Sum_Existing, SCcE1, Sum_nameplate)
    util_AddCC2 = utilizationM3 (Sum_Existing, SCcE2, Sum_nameplate)
    util_AddCC3 = utilizationM3 (Sum_Existing, SCcE3, Sum_nameplate)
    util_AddCT = utilizationM3 (Sum_Existing, SCtE, Sum_nameplate)
    util_AddCT1 = utilizationM3 (Sum_Existing, SCtE1, Sum_nameplate)
    util_AddCT2 = utilizationM3 (Sum_Existing, SCtE2, Sum_nameplate)
    util_AddCT3 = utilizationM3 (Sum_Existing, SCtE3, Sum_nameplate)
    util_AddRIC = utilizationM3 (Sum_Existing, SRicE, Sum_nameplate)
    util_AddRIC1 = utilizationM3 (Sum_Existing, SRicE1, Sum_nameplate)
    util_AddRIC2 = utilizationM3 (Sum_Existing, SRicE2, Sum_nameplate)
    util_AddRIC3 = utilizationM3 (Sum_Existing, SRicE3, Sum_nameplate)
    util_AddARO = utilizationM3 (Sum_Existing, S_aE, Sum_nameplate)
    util_AddARO1 = utilizationM3 (Sum_Existing, S_aE1, Sum_nameplate)
    util_AddARO2 = utilizationM3 (Sum_Existing, S_aE2, Sum_nameplate)
    util_AddARO3 = utilizationM3 (Sum_Existing, S_aE3, Sum_nameplate)
    # //////////////////////////////////\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\    
        
    # ==== Energy_resource, Energy_curtail, Cost_$/MWh, Utilization -Lists ====
   
    Res_list = ["Sol     ","Sol1    ","Sol2    ",
                "CC      ","CC1     ","CC2     ","CT      ","CT1     ","CT2     ",
                "RIC     ","RIC1    ","RIC2    ","ARO     ","ARO1    ","ARO2    ",
                "Bat     ","Bat1    ","Bat2    ",
                "SolCc   ","Sol1Cc  ","SolCc1  ","Sol1Cc2 ",
                "SolCt   ","Sol1Ct  ","SolCt1  ","Sol1Ct2 ",
                "SolRic  ","Sol1Ric ","SolRic1 ","Sol1Ric2",
                "Sol_a   ","Sol1_a  ","Sol_a1  ","Sol1_a2 "]
    
    Size1_list = [Add_sz,SZadd1 ,SZadd2 ,
                CC_sz,SZCC1,SZCC2,CT_sz,SZCT1,SZCT2,
                RIC_sz,SZRIC1,SZRIC2,ARO_sz,SZARO1,SZARO2,
                AddBat1_sz*chs_Bat1,AddBat2_sz*chs_Bat2,AddBat3_sz*chs_Bat3,           
     Add_sz*(( wind_Id | solar_Id)*CC_Id), SZadd1*(chs_add1*CC_Id), 
     Add_sz*(( wind_Id | solar_Id)*chs_CC1), SZadd1*(chs_add1 *chs_CC2),
     Add_sz*(( wind_Id | solar_Id)*CT_Id), SZadd1*(chs_add1*CT_Id), 
     Add_sz*(( wind_Id | solar_Id)*chs_CT1), SZadd1*(chs_add1 *chs_CT2),
     Add_sz*(( wind_Id | solar_Id)*RIC_Id), SZadd1*(chs_add1*RIC_Id),
     Add_sz*(( wind_Id | solar_Id)*chs_RIC1),SZadd1*(chs_add1 *chs_RIC2),
     Add_sz*(( wind_Id | solar_Id)*ARO_Id),SZadd1*(chs_add1*ARO_Id),
     Add_sz*(( wind_Id | solar_Id)*chs_ARO1),SZadd1*(chs_add1 *chs_ARO2)]
    
    Size2_list = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,SZbat1,SZbat2,SZbat3,
                Add_CC_sz*(( wind_Id | solar_Id)*CC_Id), Add1_CC_sz*(chs_add1*CC_Id),
                SZaddCC2*(( wind_Id | solar_Id)*chs_CC1), SZaddCC3*(chs_add1 *chs_CC2),
                Add_CT_sz*(( wind_Id | solar_Id)*CT_Id), Add1_CT_sz*(chs_add1*CT_Id),
                SZaddCT2*(( wind_Id | solar_Id)*chs_CT1) ,SZaddCT3*(chs_add1 *chs_CT2),
                Add_RIC_sz*(( wind_Id | solar_Id)*RIC_Id), Add1_RIC_sz*(chs_add1*RIC_Id),
                SZaddRIC2*(( wind_Id | solar_Id)*chs_RIC1),SZaddRIC3*(chs_add1 *chs_RIC2),
                Add_ARO_sz*(( wind_Id | solar_Id)*ARO_Id), Add1_ARO_sz*(chs_add1*ARO_Id),
                SZaddARO2*(( wind_Id | solar_Id)*chs_ARO1),SZaddARO3*(chs_add1 *chs_ARO2)]
    
    Energy_list = [MWhS, MWhS1, MWhS2,
                   SumMWhCC,SumMWhCC1,SumMWhCC2,
                   SumMWhCT,SumMWhCT1,SumMWhCT2,
                   SumMWhRIC,SumMWhRIC1,SumMWhRIC2,
                   SumMWhA,SumMWhA1,SumMWhA2,
                   MWhBat1,MWhBat2,MWhBat3,
                   SCcE,SCcE1,SCcE2,SCcE3,SCtE,SCtE1,SCtE2,SCtE3,
                   SRicE,SRicE1,SRicE2,SRicE3,S_aE,S_aE1,S_aE2,S_aE3] 
    
    Curtail_list =[MWhS_Tcrt,MWhS1_Tcrt,MWhS2_Tcrt,
                   CC_Tcrt,CC1_Tcrt,CC2_Tcrt,
                   CT_Tcrt,CT1_Tcrt,CT2_Tcrt,
                   RIC_Tcrt,RIC1_Tcrt,RIC2_Tcrt,
                   A_Tcrt,A1_Tcrt,A2_Tcrt,
                   Bat1_Tcrt,Bat2_Tcrt,Bat3_Tcrt,
                   SCcCrE,SCcCrE1,SCcCrE2,SCcCrE3,SCtCrE,SCtCrE1,SCtCrE2,SCtCrE3,
                   SRicCrE,SRicCrE1,SRicCrE2,SRicCrE3,S_aCrE,S_aCrE1,S_aCrE2,S_aCrE3]
    
    DpMWh_list =[Sol_DpE,Sol1_DpE,Sol2_DpE,
                   CC_DpE,CC1_DpE,CC2_DpE,
                   CT_DpE,CT1_DpE,CT2_DpE,
                   RIC_DpE,RIC1_DpE,RIC2_DpE,
                   A_DpE,A1_DpE,A2_DpE,
                   Bat1_DpE,Bat2_DpE,Bat3_DpE,
                   SCcDpE,SCcDpE1,SCcDpE2,SCcDpE3,SCtDpE,SCtDpE1,SCtDpE2,SCtDpE3,
                   SRicDpE,SRicDpE1,SRicDpE2,SRicDpE3,S_aDpE,S_aDpE1,S_aDpE2,S_aDpE3]
    
    Util_list =[util_Add, util_Add1, util_Add2, util_CC, util_CC1, util_CC2,       
        util_CT,util_CT1,util_CT2, util_RIC,util_RIC1,util_RIC2, 
        util_ARO, util_ARO1, util_ARO2, util_Bat1, util_Bat2, util_Bat3,
        util_AddCC, util_AddCC1, util_AddCC2, util_AddCC3,
        util_AddCT, util_AddCT1, util_AddCT2, util_AddCT3,
        util_AddRIC, util_AddRIC1, util_AddRIC2, util_AddRIC3,
        util_AddARO, util_AddARO1, util_AddARO2, util_AddARO3]
    
    # Create a list of list named 'table [[];[];[]....]
    table = []
    row = []
    for i in range(len(Res_list)):
        row= Res_list[i], int(Energy_list[i]), int\
            (Curtail_list[i]), int(DpMWh_list [i]), \
                int(Util_list[i]),int(Size1_list[i]), int(Size2_list[i])
        table.append (row)
        #print ("\nrow",row )
    #print ("\ntable",table )
    
    # formating each line of the table
    line = \
        "| {res:<10s} | {Enrg:7d} | {Curt:7d} | {Cost:7d} | {Util:3d}|\
    {Size1:3d}|    {Size2:3d}|".format
    
    # print table header
    print ("\n""          ","  Enrg_Res"," Enrg_Curt"," $/MWh ",\
           " Utiliz "," Size1  "," Size2  ")
    # print table
    for L in table:
        print(line (res= L[0], Enrg= L[1], Curt=L[2],Cost=L[3], Util=L[4],\
                    Size1=L[5],Size2=L[6]))
        
     
    # # ((((((((((((((((((((( Plots ))))))))))))))))))))))))))))))))))))))))))))
    #---------------------- matplotlib -----------------------------------
    
    # ============ Energy plots ==============================================
    
    # 1) plot MWh as recomended by the code- IDEAL
    plt.figure (1)
    
    names = ['Solar','CC','CT','RIC','Aro',
          'Solar\n + \nBat','Solar\n + \nCc','Solar\n + \nCt',
          'Solar\n + \nRic','Solar\n + \n_aro']
    values= [MWhS,SumMWhCC,SumMWhCT,SumMWhRIC,SumMWhA,
              MWhBat2,SCcE,SCtE,SRicE,S_aE]
    
    plt.title('Energy by Resources')
    plt.ylabel('Energy [MWh]')
    plt.xlabel('Resources')
    plt.bar(names, values, color='b')
    
    plt.show()
    
    # 2) plot Curtail_MWh as recomended by the code -IDEAL
    plt.figure (2)
    
    names = ['Solar','CC','CT','RIC','Aro',
          'Solar\n + \nBat','Solar\n + \nCc','Solar\n + \nCt',
          'Solar\n + \nRic','Solar\n + \n_aro']
    values= [MWhS_Tcrt,CC_Tcrt,CT_Tcrt,RIC_Tcrt,A_Tcrt,
             Bat2_Tcrt,SCcCrE,SCtCrE,SRicCrE,S_aCrE]
    
    plt.title('Curtail Energy by Resources')
    plt.ylabel('Curtail Energy [MWh]')
    plt.xlabel('Resources')
    plt.bar(names, values, color='b')
    
    plt.show()
    
    # ============ $/MWh plots ================================================
    
    # 3) plot 'Energy Price per Resource' for all IDEAL resources
    plt.figure (3)
    
    names = ['Solar','CC','CT','RIC','Aro',
          'Solar\n + \nBat','Solar\n + \nCc','Solar\n + \nCt',
          'Solar\n + \nRic','Solar\n + \n_aro']
    values= [Sol_DpE,CC_DpE,CT_DpE,RIC_DpE,A_DpE,
          Bat2_DpE,SCcDpE,SCtDpE,SRicDpE,S_aDpE]
    
    plt.title('Energy Price per Resource')
    plt.ylabel('dollars per Energy [$/MWh]')
    plt.xlabel('Resources')
    plt.bar(names, values,color ='c')
    
    plt.show()
    
    # 4) plot 'Energy Price per Resource' for all SINGLE resources
    plt.figure (4)
    
    names = ['Solar','Solar1','Solar2',
             'CC','CC1','CC2','CT','CT1','CT2',
            'RIC', 'RIC1','RIC2','Aro','Aro1','Aro2']
    values= [Sol_DpE,Sol1_DpE,Sol2_DpE,
             CC_DpE,CC1_DpE,CC2_DpE,
             CT_DpE, CT1_DpE, CT2_DpE,
             RIC_DpE,RIC1_DpE,RIC2_DpE,
             A_DpE,A1_DpE,A2_DpE]
    
    plt.title('Energy Price per Single Resources')
    plt.ylabel('dollars per Energy [$/MWh]')
    plt.xlabel('Single Resources')
    plt.bar(names, values,color ='c')
    
    plt.show()
    
    # 5) plot 'Energy Price per Resource' for Combinational resources
    plt.figure (5)
    
    names = ['Solar\n + \nBat','Solar\n + \nBat1','Solar\n + \nBat2',
             'Solar\n + \nCc', 'Solar1\n + \nCc', 'Solar\n + \nCc1', 'Solar1\n + \nCc2',
             'Solar\n + \nCt', 'Solar1\n + \nCt', 'Solar\n + \nCt1', 'Solar1\n + \nCt2']
    values= [Bat1_DpE,Bat2_DpE,Bat3_DpE,
             SCcDpE,SCcDpE1,SCcDpE2,SCcDpE3,
             SCtDpE,SCtDpE1,SCtDpE2,SCtDpE3 ]
    
    plt.title('Energy Price per Combinational Resources')
    plt.ylabel('dollars per Energy [$/MWh]')
    plt.xlabel('Combinational Resources')
    plt.bar(names, values,color ='c')
    
    plt.show()
    
    # ============ Utilization plots ==========================================
    
    # 6) plot 'Utilization per Resource' IDEAL
    plt.figure (6)
    
    names = ['Solar','CC','CT','RIC','Aro',
          'Solar\n + \nBat','Solar\n + \nCc','Solar\n + \nCt',
          'Solar\n + \nRic','Solar\n + \n_aro']
    values= [ util_Add, util_CC, util_CT, util_RIC, util_ARO,
           util_Bat2,util_AddCC, util_AddCT, util_AddRIC, util_AddARO]
    
    plt.title('Energy Utilization per Resource')
    plt.ylabel('Utilization [fract.]')
    plt.xlabel('Resources')
    plt.bar(names, values,color ='r')
    
    plt.show()
    
    # 7) plot 'Energy Price per Resource' for all SINGLE resources
    plt.figure (7)
    
    names = ['Solar','Solar1','Solar2',
             'CC','CC1','CC2','CT','CT1','CT2',
            'RIC', 'RIC1','RIC2','Aro','Aro1','Aro2']
    values= [ util_Add, util_Add1, util_Add2, util_CC, util_CC1, util_CC2,       
        util_CT,util_CT1,util_CT2, util_RIC,util_RIC1,util_RIC2,
        util_ARO, util_ARO1, util_ARO2]
    
    plt.title('Energy Utilization per Single Resources')
    plt.ylabel('Utilization [fract.]')
    plt.xlabel('Single Resources')
    plt.bar(names, values,color ='r')
    
    plt.show()
      
    # 8) plot 'Energy Price per Resource' for Combinational resources
    plt.figure (8)
    
    names = ['Solar\n + \nBat','Solar\n + \nBat1','Solar\n + \nBat2',
             'Solar\n + \nCc', 'Solar1\n + \nCc', 'Solar\n + \nCc1', 'Solar1\n + \nCc2',
             'Solar\n + \nCt', 'Solar1\n + \nCt', 'Solar\n + \nCt1', 'Solar1\n + \nCt2']
    values= [ util_Bat1, util_Bat2, util_Bat3,
        util_AddCC, util_AddCC1, util_AddCC2, util_AddCC3,
        util_AddCT, util_AddCT1, util_AddCT2, util_AddCT3,]
    
    plt.title('Energy Utilization per Combinational Resources')
    plt.ylabel('Utilization [fract.]')
    plt.xlabel('Combinational Resources')
    plt.bar(names, values,color ='r')
    
    plt.show()
    
    # *************** assign values to xlsx cells ************************                                 
    
    columnD_wind = ['D4', 'D5','D6','D7','D8','D9','D10','D11','D12','D13','D14',
                'D15','D16','D17','D18','D19','D20','D21','D22','D23','D24',
                'D25','D26','D27','D28','D29','D30','D31','D32','D33','D34',
                'D35','D36','D37']
    
    columnE_wind = ['E4', 'E5','E6','E7','E8','E9','E10','E11','E12','E13','E14',
                'E15','E16','E17','E18','E19','E20','E21','E22','E23','E24',
                'E25','E26','E27','E28','E29','E30','E31','E32','E33','E34',
                'E35','E36','E37']
    
    columnF_wind = ['F4', 'F5','F6','F7','F8','F9','F10','F11','F12','F13','F14',
                'F15','F16','F17','F18','F19','F20','F21','F22','F23','F24',
                'F25','F26','F27','F28','F29','F30','F31','F32','F33','F34',
                'F35','F36','F37']
    
    columnG_wind = ['G4', 'G5','G6','G7','G8','G9','G10','G11','G12','G13','G14',
                'G15','G16','G17','G18','G19','G20','G21','G22','G23','G24',
                'G25','G26','G27','G28','G29','G30','G31','G32','G33','G34',
                'G35','G36','G37']
    
    columnH_wind = ['H4', 'H5','H6','H7','H8','H9','H10','H11','H12','H13','H14',
                'H15','H16','H17','H18','H19','H20','H21','H22','H23','H24',
                'H25','H26','H27','H28','H29','H30','H31','H32','H33','H34',
                'H35','H36','H37']
    
    columnI_wind = ['I4', 'I5','I6','I7','I8','I9','I10','I11','I12','I13','I14',
                'I15','I16','I17','I18','I19','I20','I21','I22','I23','I24',
                'I25','I26','I27','I28','I29','I30','I31','I32','I33','I34',
                'I35','I36','I37']

    def values_to_cells (col, lists):
        for j in range(len(lists)):
            x = col[j]
            page[x] = lists[j]
        return page
    
    energy_col = values_to_cells (columnD_wind, Energy_list)
    curtail_col = values_to_cells (columnE_wind, Curtail_list)
    cost_col = values_to_cells (columnF_wind,  DpMWh_list)
    utilization_col = values_to_cells (columnG_wind, Util_list)
    size1_col = values_to_cells (columnH_wind, Size1_list)
    size2_col = values_to_cells (columnI_wind, Size2_list)
    
    # **********************************************************************
    # ------------------- excel graphs ---------------------------------------
    #--------Excel using (openpyxl)---From Existing Wind---------------------
    
    # Figure 1 'Energy Price per Resource'
    W_dolpMWh = Reference(page, min_row=4, max_row=37, min_col=6,max_col=6)
    resources = Reference (page, min_row=4, max_row=37, min_col=3,max_col=3 )
    DpEplot =Series(W_dolpMWh, title= 'resources')                             
    chartDpE = BarChart()
    chartDpE.append(DpEplot)
    chartDpE.set_categories(resources)
    # labeling
    chartDpE.title = 'Energy Price per Resource'
    chartDpE.x_axis.title = 'Resources'
    chartDpE.y_axis.title = 'Energy Cost [$/MWh]'
    # chart dimentions
    chartDpE.height = 10 # default is 7.5
    chartDpE.width = 20  # default is 15
    # place on excel
    page.add_chart(chartDpE,'k4')
    
    # Figure 2 'Utilization vs Resources'
    W_Util = Reference(page, min_row=4, max_row=37, min_col=7,max_col=7)
    resources = Reference (page, min_row=4, max_row=37, min_col=3,max_col=3 )
    Ut_plot =Series(W_Util, title= 'resources')                             
    chartW_Ut = BarChart()
    chartW_Ut.append(Ut_plot)
    chartW_Ut.set_categories(resources)
    # labeling
    chartW_Ut.title = 'Utilization vs Resources'
    chartW_Ut.x_axis.title = 'Resources'
    chartW_Ut.y_axis.title = 'Utilization [%]'
    # chart dimentions
    chartW_Ut.height = 10 # default is 7.5
    chartW_Ut.width = 20  # default is 15
    # place on excel
    page.add_chart(chartW_Ut,'k30')
    
    #-------- Excel graphing Ends ------------------------------------------

# ############## Existing Solar Starts #######################################      
elif loc == Existing_Solar: # the Additional resource is wind
    
    # the 3 Additional resource senarios
    # (CpE, MWhAddt, MWhBatt, Tcrt):
    MWhW , MWhW_Tcrt, Wind_D, WindCurt_D, W_DpE = AdditionalBatTF \
        (C26, MWhAddtn, MWhAddtn_Tcrt)    # ideal Additional  
    MWhW1, MWhW1_Tcrt, Wind1_D, WindCurt1_D, W1_DpE = AdditionalBatTF \
        (C26, MWhAddtn1, MWhAddtn1_Tcrt)  # 1. cho1 Additional 
    MWhW2, MWhW2_Tcrt, Wind2_D, WindCurt2_D, W2_DpE = AdditionalBatTF \
        (C26, MWhAddtn2, MWhAddtn2_Tcrt)  # 2. cho2 Additional
    #print ("\nMWhW: ", MWhW, "MWhW_Tcrt: ", MWhW_Tcrt, "Wind_D: ", Wind_D, \
    #       "WindCurt_D: ",WindCurt_D, "W_DpE: ",W_DpE )
     
    # Comb. Battery 
    # 3 choice user picks battery, code picks ideal Additional size
    MWhBat1, Bat1_Tcrt, Bat1_D, Bat1Curt_D, Bat1_DpE = AdditionalBatTF \
        (C26, MWhAddtn1_bat, MWhAddtnBat1_Tcrt )    # ideal Additional  
    MWhBat2, Bat2_Tcrt, Bat2_D, Bat2Curt_D, Bat2_DpE = AdditionalBatTF \
        (C26, MWhAddtn2_bat, MWhAddtnBat2_Tcrt )  # 1. cho1 Additional 
    MWhBat3, Bat3_Tcrt, Bat3_D, Bat3Curt_D, Bat3_DpE = AdditionalBatTF \
        (C26, MWhAddtn3_bat, MWhAddtnBat3_Tcrt )  # 2. cho2 Additional
    
    # the 4 comb. (Additional resource + gas) senarios
    # gasCombResource (CpE, HR,VOM,frw_crv, MWhAddt, MWh_gas, Tcrt)
     
    # CC
    WCcE, WCcCrE, WCcD, WCcCrD, WCcDpE = gasCombResource \
        (C26, CC_HR,CC_VOM,frw_crv, MWhAddtn, MWhCc, MWhAddtn_Tcrt)    # both ideal CC 
    WCcE1, WCcCrE1, WCcD1, WCcCrD1, WCcDpE1 = gasCombResource \
        (C26, CC_HR,CC_VOM,frw_crv, MWhAddtn1, MWhCc1, MWhAddtn1_Tcrt) # 1. cho1, ideal CC 
    WCcE2, WCcCrE2, WCcD2, WCcCrD2, WCcDpE2 = gasCombResource \
        (C26, CC_HR,CC_VOM,frw_crv, MWhAddtn, MWhCc2, MWhAddtn_Tcrt)   # 2. ideal, cho1 CC 
    WCcE3, WCcCrE3, WCcD3, WCcCrD3, WCcDpE3 = gasCombResource \
        (C26, CC_HR,CC_VOM,frw_crv, MWhAddtn1, MWhCc3, MWhAddtn1_Tcrt) # 3. cho1, cho2 CC 
    #print ("\nWCcE: ", WCcE, "WCcCrE: ", WCcCrE, "WCcD: ", WCcD, \
    #       "WCcCrD: ",WCcCrD, "WCcDpE: ",WCcDpE ) 
    
    # CT
    WCtE, WCtCrE, WCtD, WCtCrD, WCtDpE = gasCombResource \
        (C26, CT_HR,CT_VOM,frw_crv, MWhAddtn, MWhCt, MWhAddtn_Tcrt)   # both ideal CT 
    WCtE1, WCtCrE1, WCtD1, WCtCrD1, WCtDpE1 = gasCombResource \
        (C26, CT_HR,CT_VOM,frw_crv, MWhAddtn1, MWhCt1, MWhAddtn1_Tcrt)# 1. cho1, ideal CT 
    WCtE2, WCtCrE2, WCtD2, WCtCrD2, WCtDpE2 = gasCombResource \
        (C26, CT_HR,CT_VOM,frw_crv, MWhAddtn, MWhCt2, MWhAddtn_Tcrt)  # 2. ideal, cho1 CT
    WCtE3, WCtCrE3, WCtD3, WCtCrD3, WCtDpE3 = gasCombResource \
        (C26, CT_HR,CT_VOM,frw_crv, MWhAddtn1, MWhCt3, MWhAddtn1_Tcrt)# 3. cho1, cho2 CT 
    
    # RIC 
    WRicE, WRicCrE, WRicD, WRicCrD, WRicDpE = gasCombResource \
        (C26, RIC_HR,RIC_VOM,frw_crv, MWhAddtn, MWhRic, MWhAddtn_Tcrt)   # both ideal RIC 
    WRicE1, WRicCrE1, WRicD1, WRicCrD1, WRicDpE1 = gasCombResource \
        (C26, RIC_HR,RIC_VOM,frw_crv, MWhAddtn1, MWhRic1, MWhAddtn1_Tcrt)# 1. cho1, ideal RIC 
    WRicE2, WRicCrE2, WRicD2, WRicCrD2, WRicDpE2 = gasCombResource \
        (C26, RIC_HR,RIC_VOM,frw_crv, MWhAddtn, MWhRic2, MWhAddtn_Tcrt)  # 2. ideal, cho1 RIC 
    WRicE3, WRicCrE3, WRicD3, WRicCrD3, WRicDpE3 = gasCombResource \
        (C26, RIC_HR,RIC_VOM,frw_crv, MWhAddtn1, MWhRic3, MWhAddtn1_Tcrt)  # 3. cho1, cho2 RIC 
    
    # ARO or A _a
    W_aE, W_aCrE, W_aD, W_aCrD, W_aDpE = gasCombResource \
        (C26, ARO_HR,ARO_VOM,frw_crv, MWhAddtn, MWh_a, MWhAddtn_Tcrt)   # both ideal ARO 
    W_aE1, W_aCrE1, W_aD1, W_aCrD1, W_aDpE1 = gasCombResource \
        (C26, ARO_HR,ARO_VOM,frw_crv, MWhAddtn1, MWh_a1, MWhAddtn1_Tcrt)# 1. cho1, ideal ARO 
    W_aE2, W_aCrE2, W_aD2, W_aCrD2, W_aDpE2 = gasCombResource \
        (C26, ARO_HR,ARO_VOM,frw_crv, MWhAddtn, MWh_a2, MWhAddtn_Tcrt)  # 2. ideal, cho1 ARO 
    W_aE3, W_aCrE3, W_aD3, W_aCrD3, W_aDpE3 = gasCombResource \
        (C26, ARO_HR,ARO_VOM,frw_crv, MWhAddtn1, MWh_a3, MWhAddtn1_Tcrt)  # 3. cho1, cho2 ARO   
    
    # ===========================================================================
    # ///////////// test for percentage of line utilization \\\\\\\\\\\\\\\\\\\\\
    # finds line utilization as a %
    # Additional (alone or in combination) resources (depends on existing resource)
    util_Add = utilizationM3 (Sum_Existing, MWhW, Sum_nameplate)
    util_Add1 = utilizationM3 (Sum_Existing, MWhW1, Sum_nameplate)
    util_Add2 = utilizationM3 (Sum_Existing, MWhW2, Sum_nameplate)
    util_Bat1 = utilizationM3 (Sum_Existing, MWhBat1, Sum_nameplate)
    util_Bat2 = utilizationM3 (Sum_Existing, MWhBat2, Sum_nameplate)
    util_Bat3 = utilizationM3 (Sum_Existing, MWhBat3, Sum_nameplate)
    util_AddCC = utilizationM3 (Sum_Existing, WCcE, Sum_nameplate)
    util_AddCC1 = utilizationM3 (Sum_Existing, WCcE1, Sum_nameplate)
    util_AddCC2 = utilizationM3 (Sum_Existing, WCcE2, Sum_nameplate)
    util_AddCC3 = utilizationM3 (Sum_Existing, WCcE3, Sum_nameplate)
    util_AddCT = utilizationM3 (Sum_Existing, WCtE, Sum_nameplate)
    util_AddCT1 = utilizationM3 (Sum_Existing, WCtE1, Sum_nameplate)
    util_AddCT2 = utilizationM3 (Sum_Existing, WCtE2, Sum_nameplate)
    util_AddCT3 = utilizationM3 (Sum_Existing, WCtE3, Sum_nameplate)
    util_AddRIC = utilizationM3 (Sum_Existing, WRicE, Sum_nameplate)
    util_AddRIC1 = utilizationM3 (Sum_Existing, WRicE1, Sum_nameplate)
    util_AddRIC2 = utilizationM3 (Sum_Existing, WRicE2, Sum_nameplate)
    util_AddRIC3 = utilizationM3 (Sum_Existing, WRicE3, Sum_nameplate)
    util_AddARO = utilizationM3 (Sum_Existing, W_aE, Sum_nameplate)
    util_AddARO1 = utilizationM3 (Sum_Existing, W_aE1, Sum_nameplate)
    util_AddARO2 = utilizationM3 (Sum_Existing, W_aE2, Sum_nameplate)
    util_AddARO3 = utilizationM3 (Sum_Existing, W_aE3, Sum_nameplate)
    # //////////////////////////////////\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\    
        
    
    # ==== Energy_resource, Energy_curtail, Cost_$/MWh, Utilization -Lists ====
    Res_list = ["Wind    ","Wind1   ","Wind2   ",
                "CC      ","CC1     ","CC2     ","CT      ","CT1     ","CT2     ",
                "RIC     ","RIC1    ","RIC2    ","ARO     ","ARO1    ","ARO2    ",
                "Bat     ","Bat1    ","Bat2    ",
                "WindCc  ","Wind1Cc ","WindCc1 ","Wind1Cc2 ",
                "WindCt  ","Wind1Ct ","WindCt1 ","Wind1Ct2 ",
                "WindRic ","Wind1Ri ","WindRic1","Wind1Ric2",
                "Wind_a  ","Wind1_a ","Wind_a1 ","Wind1_a2 "]
    
    Size1_list = [Add_sz,SZadd1 ,SZadd2 ,
                CC_sz,SZCC1,SZCC2,CT_sz,SZCT1,SZCT2,
                RIC_sz,SZRIC1,SZRIC2,ARO_sz,SZARO1,SZARO2,
                AddBat1_sz*chs_Bat1,AddBat2_sz*chs_Bat2,AddBat3_sz*chs_Bat3,           
     Add_sz*(( wind_Id | solar_Id)*CC_Id), SZadd1*(chs_add1*CC_Id), 
     Add_sz*(( wind_Id | solar_Id)*chs_CC1), SZadd1*(chs_add1 *chs_CC2),
     Add_sz*(( wind_Id | solar_Id)*CT_Id), SZadd1*(chs_add1*CT_Id), 
     Add_sz*(( wind_Id | solar_Id)*chs_CT1), SZadd1*(chs_add1 *chs_CT2),
     Add_sz*(( wind_Id | solar_Id)*RIC_Id), SZadd1*(chs_add1*RIC_Id),
     Add_sz*(( wind_Id | solar_Id)*chs_RIC1),SZadd1*(chs_add1 *chs_RIC2),
     Add_sz*(( wind_Id | solar_Id)*ARO_Id),SZadd1*(chs_add1*ARO_Id),
     Add_sz*(( wind_Id | solar_Id)*chs_ARO1),SZadd1*(chs_add1 *chs_ARO2)]
    
    Size2_list = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,SZbat1,SZbat2,SZbat3,
                Add_CC_sz*(( wind_Id | solar_Id)*CC_Id), Add1_CC_sz*(chs_add1*CC_Id),
                SZaddCC2*(( wind_Id | solar_Id)*chs_CC1), SZaddCC3*(chs_add1 *chs_CC2),
                Add_CT_sz*(( wind_Id | solar_Id)*CT_Id), Add1_CT_sz*(chs_add1*CT_Id),
                SZaddCT2*(( wind_Id | solar_Id)*chs_CT1) ,SZaddCT3*(chs_add1 *chs_CT2),
                Add_RIC_sz*(( wind_Id | solar_Id)*RIC_Id), Add1_RIC_sz*(chs_add1*RIC_Id),
                SZaddRIC2*(( wind_Id | solar_Id)*chs_RIC1),SZaddRIC3*(chs_add1 *chs_RIC2),
                Add_ARO_sz*(( wind_Id | solar_Id)*ARO_Id), Add1_ARO_sz*(chs_add1*ARO_Id),
                SZaddARO2*(( wind_Id | solar_Id)*chs_ARO1),SZaddARO3*(chs_add1 *chs_ARO2)]
    
    Energy_list = [MWhW, MWhW1, MWhW2,
                   SumMWhCC,SumMWhCC1,SumMWhCC2,
                   SumMWhCT,SumMWhCT1,SumMWhCT2,
                   SumMWhRIC,SumMWhRIC1,SumMWhRIC2,
                   SumMWhA,SumMWhA1,SumMWhA2,
                   MWhBat1,MWhBat2,MWhBat3,
                   WCcE,WCcE1,WCcE2,WCcE3,WCtE,WCtE1,WCtE2,WCtE3,
                   WRicE,WRicE1,WRicE2,WRicE3,W_aE,W_aE1,W_aE2,W_aE3] 
    
    Curtail_list =[MWhW_Tcrt,MWhW1_Tcrt,MWhW2_Tcrt,
                   CC_Tcrt,CC1_Tcrt,CC2_Tcrt,
                   CT_Tcrt,CT1_Tcrt,CT2_Tcrt,
                   RIC_Tcrt,RIC1_Tcrt,RIC2_Tcrt,
                   A_Tcrt,A1_Tcrt,A2_Tcrt,
                   Bat1_Tcrt,Bat2_Tcrt,Bat3_Tcrt,
                   WCcCrE,WCcCrE1,WCcCrE2,WCcCrE3,WCtCrE,WCtCrE1,WCtCrE2,WCtCrE3,
                   WRicCrE,WRicCrE1,WRicCrE2,WRicCrE3,W_aCrE,W_aCrE1,W_aCrE2,W_aCrE3]
    
    DpMWh_list =[W_DpE,W1_DpE,W2_DpE,
                   CC_DpE,CC1_DpE,CC2_DpE,
                   CT_DpE,CT1_DpE,CT2_DpE,
                   RIC_DpE,RIC1_DpE,RIC2_DpE,
                   A_DpE,A1_DpE,A2_DpE,
                   Bat1_DpE,Bat2_DpE,Bat3_DpE,
                   WCcDpE,WCcDpE1,WCcDpE2,WCcDpE3,WCtDpE,WCtDpE1,WCtDpE2,WCtDpE3,
                   WRicDpE,WRicDpE1,WRicDpE2,WRicDpE3,W_aDpE,W_aDpE1,W_aDpE2,W_aDpE3]
    
    Util_list =[util_Add, util_Add1, util_Add2, util_CC, util_CC1, util_CC2,       
        util_CT,util_CT1,util_CT2, util_RIC,util_RIC1,util_RIC2,
        util_ARO, util_ARO1, util_ARO2, util_Bat1, util_Bat2, util_Bat3,
        util_AddCC, util_AddCC1, util_AddCC2, util_AddCC3,
        util_AddCT, util_AddCT1, util_AddCT2, util_AddCT3,
        util_AddRIC, util_AddRIC1, util_AddRIC2, util_AddRIC3,
        util_AddARO, util_AddARO1, util_AddARO2, util_AddARO3]
    
    # Create a list of list named 'table [[];[];[]....]
    table = []
    row = []
    for i in range(len(Res_list)):
        row= Res_list[i], int(Energy_list[i]), int\
            (Curtail_list[i]), int(DpMWh_list [i]), \
                int(Util_list[i]),int(Size1_list[i]), int(Size2_list[i])
        table.append (row)
        #print ("\nrow",row )
    #print ("\ntable",table )
    
    # formating each line of the table
    line = \
        "| {res:<10s} | {Enrg:7d} | {Curt:7d} | {Cost:7d} | {Util:3d}|\
    {Size1:3d}|    {Size2:3d}|".format
    
    # print table header
    print ("\n""          ","  Enrg_Res"," Enrg_Curt"," $/MWh ",\
           " Utiliz "," Size1  "," Size2  ")
    # print table
    for L in table:
        print(line (res= L[0], Enrg= L[1], Curt=L[2],Cost=L[3], Util=L[4],\
                    Size1=L[5],Size2=L[6]))
        
    # # ((((((((((((((((((((( Plots ))))))))))))))))))))))))))))))))))))))))))))
    #---------------------- matplotlib -----------------------------------
    # ============ Energy plots ==============================================
     
    # 1) plot MWh as recomended by the code - IDEAL
    plt.figure (1)
    
    names = ['Wind','CC','CT','RIC','Aro',
          'Wind\n + \nBat','Wind\n + \nCc','Wind\n + \nCt',
          'Wind\n + \nRic','Wind\n + \n_aro']
    values= [MWhW,SumMWhCC,SumMWhCT,SumMWhRIC,SumMWhA,
              MWhBat2,WCcE,WCtE,WRicE,W_aE]
    
    plt.title('Energy by Resources')
    plt.ylabel('Energy [MWh]')
    plt.xlabel('Resources')
    plt.bar(names, values, color='b')
    
    plt.show()
    
    # 2) plot Curtail_MWh as recomended by the code - IDEAL
    plt.figure (2)
    
    names = ['Wind','CC','CT','RIC','Aro',
          'Wind\n + \nBat','Wind\n + \nCc','Wind\n + \nCt',
          'Wind\n + \nRic','Wind\n + \n_aro']
    values= [MWhW_Tcrt,CC_Tcrt,CT_Tcrt,RIC_Tcrt,A_Tcrt,
             Bat2_Tcrt,WCcCrE,WCtCrE,WRicCrE,W_aCrE]
    
    plt.title('Curtail Energy by Resources')
    plt.ylabel('Curtail Energy [MWh]')
    plt.xlabel('Resources')
    plt.bar(names, values, color='b')
    
    plt.show()
    
    # ============ $/MWh plots ================================================
   
    # 3) plot 'Energy Price per Resource' for all IDEAL resources
    plt.figure (3)
    
    names = ['Wind','CC','CT','RIC','Aro',
          'Wind\n + \nBat','Wind\n + \nCc','Wind\n + \nCt',
          'Wind\n + \nRic','Wind\n + \n_aro']
    values= [W_DpE,CC_DpE,CT_DpE,RIC_DpE,A_DpE,
          Bat2_DpE,WCcDpE,WCtDpE,WRicDpE,W_aDpE]
    
    plt.title('Energy Price per Resource')
    plt.ylabel('dollars per Energy [$/MWh]')
    plt.xlabel('Resources')
    plt.bar(names, values,color ='c')
    
    plt.show()
    
    # 4) plot 'Energy Price per Resource' for all SINGLE resources
    plt.figure (4)
    
    names = ['Wind','Wind1','Wind2',
             'CC','CC1','CC2','CT','CT1','CT2',
            'RIC', 'RIC1','RIC2','Aro','Aro1','Aro2']
    values= [W_DpE,W1_DpE,W2_DpE,
             CC_DpE,CC1_DpE,CC2_DpE,
             CT_DpE, CT1_DpE, CT2_DpE,
             RIC_DpE,RIC1_DpE,RIC2_DpE,
             A_DpE,A1_DpE,A2_DpE]
    
    plt.title('Energy Price per Single Resources')
    plt.ylabel('dollars per Energy [$/MWh]')
    plt.xlabel('Resources')
    plt.bar(names, values,color ='c')
    
    plt.show()
     
     # 5) plot 'Energy Price per Resource' for Combinational resources
    plt.figure (5)
    
    names = ['Wind\n + \nBat','Wind\n + \nBat1','Wind\n + \nBat2',
             'Wind\n + \nCc', 'Wind1\n + \nCc', 'Wind\n + \nCc1', 'Wind1\n + \nCc2',
             'Wind\n + \nCt', 'Wind1\n + \nCt', 'Wind\n + \nCt1', 'Wind1\n + \nCt2']
    values= [Bat1_DpE,Bat2_DpE,Bat3_DpE,
             WCcDpE,WCcDpE1,WCcDpE2,WCcDpE3,
             WCtDpE,WCtDpE1,WCtDpE2,WCtDpE3 ]
    
    plt.title('Energy Price per Combinational Resources')
    plt.ylabel('dollars per Energy [$/MWh]')
    plt.xlabel('Combinational Resources')
    plt.bar(names, values,color ='c')
    
    plt.show()
    
    # ============ Utilization plots ==========================================
    
    # 6) plot 'Utilization per Resource' IDEAL resources
    plt.figure (6)
    
    names = ['Wind','CC','CT','RIC','Aro',
          'Wind\n + \nBat','Wind\n + \nCc','Wind\n + \nCt',
          'Wind\n + \nRic','Wind\n + \n_aro']
    values= [  util_Add, util_CC, util_CT, util_RIC, util_ARO,
           util_Bat2,util_AddCC, util_AddCT, util_AddRIC, util_AddARO]
    
    plt.title('Energy Utilization per Resource')
    plt.ylabel('Utilization [fract.]')
    plt.xlabel('Resources')
    plt.bar(names, values,color ='r')
    
    plt.show()
    
    # 7) plot 'Urilization per Resource' for all SINGLE resources
    plt.figure (7)
    
    names = ['Wind','Wind1','Wind2',
             'CC','CC1','CC2','CT','CT1','CT2',
            'RIC', 'RIC1','RIC2','Aro','Aro1','Aro2']
    values= [ util_Add, util_Add1, util_Add2, util_CC, util_CC1, util_CC2,       
        util_CT,util_CT1,util_CT2, util_RIC,util_RIC1,util_RIC2,
        util_ARO, util_ARO1, util_ARO2]
    
    plt.title('Energy Utilization per Single Resources')
    plt.ylabel('Utilization [fract.]')
    plt.xlabel('Resources')
    plt.bar(names, values,color ='r')
    
    plt.show()
     
    # 8) plot 'Energy Price per Resource' for Combinational resources
    plt.figure (8)
    
    names = ['Wind\n + \nBat','Wind\n + \nBat1','Wind\n + \nBat2',
             'Wind\n + \nCc', 'Wind1\n + \nCc', 'Wind\n + \nCc1', 'Wind1\n + \nCc2',
             'Wind\n + \nCt', 'Wind1\n + \nCt', 'Wind\n + \nCt1', 'Wind1\n + \nCt2']
    values= [util_Bat1, util_Bat2, util_Bat3,
        util_AddCC, util_AddCC1, util_AddCC2, util_AddCC3,
        util_AddCT, util_AddCT1, util_AddCT2, util_AddCT3,]
    
    plt.title('Energy Utilization per Combinational Resources')
    plt.ylabel('Utilization [fract.]')
    plt.xlabel('Combinational Resources')
    plt.bar(names, values,color ='r')
    
    plt.show()   

    # *************** assign values to xlsx cells ************************                                 
    
    columnD_solar = ['D47', 'D48','D49','D50','D51','D52','D53','D54',
                    'D55','D56','D57','D58','D59','D60','D61','D62','D63',
                'D64','D65','D66','D67','D68','D69','D70','D71','D72','D73',
                'D74','D75','D76','D77','D78','D79','D80']
    
    columnE_solar = ['E47', 'E48','E49','E50','E51','E52','E53','E54',
                    'E55','E56','E57','E58','E59','E60','E61','E62','E63',
                'E64','E65','E66','E67','E68','E69','E70','E71','E72','E73',
                'E74','E75','E76','E77','E78','E79','E80']
    
    columnF_solar = ['F47', 'F48','F49','F50','F51','F52','F53','F54',
                    'F55','F56','F57','F58','F59','F60','F61','F62','F63',
                'F64','F65','F66','F67','F68','F69','F70','F71','F72','F73',
                'F74','F75','F76','F77','F78','F79','F80']
    
    columnG_solar = ['G47', 'G48','G49','G50','G51','G52','G53','G54',
                    'G55','G56','G57','G58','G59','G60','G61','G62','G63',
                'G64','G65','G66','G67','G68','G69','G70','G71','G72','G73',
                'G74','G75','G76','G77','G78','G79','G80']
    
    columnH_solar = ['H47', 'H48','H49','H50','H51','H52','H53','H54',
                    'H55','H56','H57','H58','H59','H60','H61','H62','H63',
                'H64','H65','H66','H67','H68','H69','H70','H71','H72','H73',
                'H74','H75','H76','H77','H78','H79','H80']
    
    columnI_solar = ['I47', 'I48','I49','I50','I51','I52','I53','I54',
                    'I55','I56','I57','I58','I59','I60','I61','I62','I63',
                'I64','I65','I66','I67','I68','I69','I70','I71','I72','I73',
                'I74','I75','I76','I77','I78','I79','I80']

    def values_to_cells (col, lists):
        for j in range(len(lists)):
            x = col[j]
            page[x] = lists[j]
        return page
    
    energy_col = values_to_cells (columnD_solar, Energy_list)
    curtail_col = values_to_cells (columnE_solar, Curtail_list)
    cost_col = values_to_cells (columnF_solar,  DpMWh_list)
    utilization_col = values_to_cells (columnG_solar, Util_list)
    size1_col = values_to_cells (columnH_solar, Size1_list)
    size2_col = values_to_cells (columnI_solar, Size2_list)
    
    # **********************************************************************
    # ------------------- excel graphs ---------------------------------------
    #--------Excel using (openpyxl)---From Existing Solar---------------------
    
    # Figure 1 'Energy Price per Resource'
    S_dolpMWh = Reference(page, min_row=47, max_row=80, min_col=6,max_col=6)
    resources = Reference (page, min_row=47, max_row=80, min_col=3,max_col=3 )
    DpEplot =Series(S_dolpMWh, title= 'resources')                             
    chartDpE = BarChart()
    chartDpE.append(DpEplot)
    chartDpE.set_categories(resources)
    # labeling
    chartDpE.title = 'Energy Price per Resource'
    chartDpE.x_axis.title = 'Resources'
    chartDpE.y_axis.title = 'Energy Cost [$/MWh]'
    # chart dimentions
    chartDpE.height = 10 # default is 7.5
    chartDpE.width = 20  # default is 15
    # place on excel
    page.add_chart(chartDpE,'k47')
    
    # Figure 2 'Utilization vs Resources'
    S_Util = Reference(page, min_row=47, max_row=80, min_col=7,max_col=7)
    resources = Reference (page, min_row=47, max_row=80, min_col=3,max_col=3 )
    Ut_plot =Series(S_Util, title= 'resources')                             
    chartS_Ut = BarChart()
    chartS_Ut.append(Ut_plot)
    chartS_Ut.set_categories(resources)
    # labeling
    chartS_Ut.title = 'Utilization vs Resources'
    chartS_Ut.x_axis.title = 'Resources'
    chartS_Ut.y_axis.title = 'Utilization [%]'
    # chart dimentions
    chartS_Ut.height = 10 # default is 7.5
    chartS_Ut.width = 20  # default is 15
    # place on excel
    page.add_chart(chartS_Ut,'k63')
    
    #-------- Excel graphing Ends ------------------------------------------
    
# if loc= num is not found
else:
    print ("pick an integer from 0 to 6 for loc")
    
 
# ===========******'Save' 'EcoAnResulsM3.xlsx'*******=======================
wbd.save('EcoAnResulsM3.xlsx')

90

